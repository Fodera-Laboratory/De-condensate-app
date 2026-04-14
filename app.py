"""
De-condensate — Raman linescan analysis GUI
Run with:  streamlit run app.py
"""

import base64
import io
import os
import sys
import zipfile

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import analysis as an
import raman_preprocessing as rms

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="De-condensate", layout="wide", page_icon="🔬")

# ── Background image ──────────────────────────────────────────────────────────
_bg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "background.jpg")
if os.path.isfile(_bg_path):
    with open(_bg_path, "rb") as _f:
        _bg_b64 = base64.b64encode(_f.read()).decode()
    st.markdown(f"""
<style>
.stApp::before {{
    content: "";
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background-image: url("data:image/jpeg;base64,{_bg_b64}");
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    opacity: 0.4;
    z-index: 0;
    pointer-events: none;
}}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
    html, body, [class*="css"], .stMarkdown, .stText, .stMetric,
    .stSelectbox, .stRadio, .stSlider, .stNumberInput,
    .stTextInput, .stButton, .stDownloadButton, .stExpander,
    .stTabs, .stTab, .stSidebar, h1, h2, h3, h4, h5, h6, p, label {
        font-family: Arial, sans-serif !important;
    }
</style>
""", unsafe_allow_html=True)

COLORS      = ["#1b85b8", "#5a5255", "#559e83", "#ae5a41"]
COLORS_FILL = ["rgba(27,133,184,0.15)", "rgba(90,82,85,0.15)",
               "rgba(85,158,131,0.15)", "rgba(174,90,65,0.15)"]
COMP_LABELS = ["Protein", "Water", "Glass", "Salt"]

TRAINING_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "training_data")


@st.cache_data(ttl=3600)
def _fetch_pdb_ss(pdb_id):
    """Return (helix_frac, sheet_frac, other_frac) from PDBe + RCSB APIs."""
    import requests
    pid = pdb_id.strip().lower()

    # Secondary structure residue ranges from PDBe
    r_ss = requests.get(
        f"https://www.ebi.ac.uk/pdbe/api/pdb/entry/secondary_structure/{pid}",
        timeout=8,
    )
    r_ss.raise_for_status()
    ss = r_ss.json().get(pid, {})
    helix_res = sheet_res = 0
    for mol in ss.get("molecules", []):
        for chain in mol.get("chains", []):
            s = chain.get("secondary_structure", {})
            for h in s.get("helices", []):
                helix_res += h["end"]["residue_number"] - h["start"]["residue_number"] + 1
            for st in s.get("strands", []):
                sheet_res += st["end"]["residue_number"] - st["start"]["residue_number"] + 1

    # Total deposited residues from RCSB
    r_info = requests.get(
        f"https://data.rcsb.org/rest/v1/core/entry/{pid.upper()}",
        timeout=8,
    )
    r_info.raise_for_status()
    total = int(r_info.json().get("rcsb_entry_info", {}).get("deposited_polymer_monomer_count", 0))
    if total == 0:
        raise ValueError("Could not determine total residue count from RCSB.")

    h_f = helix_res / total
    s_f = sheet_res / total
    return h_f, s_f, max(0.0, 1.0 - h_f - s_f)


def _pdb_amp_priors(helix, sheet, other, n_gauss):
    """Map PDB secondary structure fractions to amplitude priors for n_gauss components.
    Component order follows Fellows et al. 2020 presets (1605, 1618, 1635, 1653, 1672, 1686).
    Beyond 6 components, flat prior is used.
    """
    # Fellows et al. mapping:
    # 1605 – sidechains/aggregates: small fixed contribution
    # 1618 – β-sheet intermolecular (aggregation): 0 for native structure
    # 1635 – β-sheet intramolecular parallel: half of sheet
    # 1653 – α-helix: helix
    # 1672 – β-turn / random coil: other
    # 1686 – β-sheet intermolecular AP + sidechains: other half of sheet
    _base6 = [0.03, 0.0, sheet * 0.5, helix, other, sheet * 0.5]
    _flat  = 1.0 / n_gauss
    _priors = [(_base6[i] if i < 6 else _flat) for i in range(n_gauss)]
    _total  = sum(_priors) or 1.0
    return [v / _total for v in _priors]


def _training_files(subdir=None):
    """
    List CSV files in training_data/ or a named subfolder (Protein / PEG / Salt).
    Returns a sorted list of filenames, or [] if folder doesn't exist.
    """
    folder = os.path.join(TRAINING_DIR, subdir) if subdir else TRAINING_DIR
    if not os.path.isdir(folder):
        return []
    return sorted(f for f in os.listdir(folder) if f.lower().endswith(".csv"))


def _training_path(subdir, filename):
    if subdir:
        return os.path.join(TRAINING_DIR, subdir, filename)
    return os.path.join(TRAINING_DIR, filename)


def _read_csv_src(src):
    """Read a CSV from either a Streamlit UploadedFile or a plain file-path string."""
    df = pd.read_csv(src)
    if hasattr(src, "seek"):   # file object — reset so it can be read again
        src.seek(0)
    return df


def _fig_caption(text: str):
    st.caption(text)


def _pls_coef(model, n_features):
    """
    Return PLSRegression.coef_ as shape (n_features, n_targets).
    sklearn < 1.1 stored it transposed as (n_targets, n_features).
    """
    coef = np.atleast_2d(model.coef_)
    if coef.shape[0] != n_features:
        coef = coef.T
    return coef


def _std_picker(label, caption, key_prefix, subdir, optional=False):
    """
    Render a standard-CSV picker with Upload / Training library (/ None) radio.
    Returns the source (UploadedFile or path string or None).
    """
    st.markdown(f"**{label}**" + ("  *(optional)*" if optional else ""))
    if caption:
        st.caption(caption)
    options = ["Upload file", "Training library"]
    if optional:
        options.append("None")
    mode = st.radio("", options, horizontal=True, key=f"{key_prefix}_mode",
                    label_visibility="collapsed")
    if mode == "Upload file":
        up = st.file_uploader("", type=["csv"], key=f"{key_prefix}_up",
                              label_visibility="collapsed")
        return up
    elif mode == "None":
        return None
    else:
        lib = _training_files(subdir)
        folder_label = subdir if subdir else "training_data"
        sel = st.selectbox(
            "", lib if lib else [f"(no files in {folder_label}/)"],
            key=f"{key_prefix}_sel", label_visibility="collapsed",
        )
        return _training_path(subdir, sel) if lib else None


# ─────────────────────────────────────────────────────────────────────────────
# Tab-navigation callbacks (must be defined before any button that uses them)
def _goto_pls():     st.session_state["active_tab"] = "📊  PLS regression"
def _goto_mcr():     st.session_state["active_tab"] = "🔬  MCR decomposition"
def _goto_further(): st.session_state["active_tab"] = "🔍  Further analysis"


def _std_spectra_plot(df, title, conc_unit, color, height=300, is_salt=False, use_pls_settings=False):
    """Plot preprocessed spectra from a standard CSV coloured by concentration."""
    try:
        wn_raw = pd.to_numeric(df.columns[1:], errors="coerce").values
        X_raw  = df.iloc[:, 1:].values.astype(float)
        c      = df.iloc[:, 0].values.astype(float)
        _s = pls_settings if use_pls_settings else settings
        with st.spinner("Preprocessing…"):
            X_proc, wn_proc, _ = an.preprocess_matrix(X_raw, wn_raw, _s,
                                                       is_salt=is_salt)
        fig = go.Figure()
        for i in range(X_proc.shape[0]):
            fig.add_trace(go.Scatter(
                x=wn_proc, y=X_proc[i], mode="lines",
                line=dict(color=color, width=0.9),
                opacity=0.55, showlegend=False,
                hovertemplate=f"{c[i]:.2f} {conc_unit}<extra></extra>",
            ))
        fig.update_layout(
            title=title, height=height,
            xaxis_title="Wavenumber (cm⁻¹)",
            yaxis_title="Intensity (a.u.)",
            margin=dict(t=40, b=40),
        )
        st.plotly_chart(fig, use_container_width=True)
        st.caption(
            f"{X_proc.shape[0]} preprocessed spectra, "
            f"{wn_proc[0]:.0f}–{wn_proc[-1]:.0f} cm⁻¹, "
            f"concentration range {c.min():.1f}–{c.max():.1f} {conc_unit}."
        )
    except Exception as _e:
        st.warning(f"Could not plot spectra: {_e}")


# ─────────────────────────────────────────────────────────────────────────────
# Tabs
# ─────────────────────────────────────────────────────────────────────────────
tab_tutorial, tab_files, tab_pls, tab_mcr, tab_further, tab_image, tab_download, tab_training, tab_about = st.tabs(
    ["📖  Tutorial", "📂  Files & preprocessing", "📊  PLS regression", "🔬  MCR decomposition",
     "🔍  Further analysis", "🗺️  Image overlay", "⬇  Download",
     "🗂  Training data", "ℹ  About"],
    default=st.session_state.get("active_tab", "📖  Tutorial"),
)
# Alias old tab variables so existing with-blocks render in the right new tabs
tab_preview  = tab_files   # data preview appears at bottom of Files tab
tab_calib    = tab_pls     # calibration results appear at bottom of PLS tab
tab_results  = tab_mcr     # linescan results appear at bottom of MCR tab

# ─────────────────────────────────────────────────────────────────────────────
# Files tab — linescan upload + global preprocessing + spectral range
# ─────────────────────────────────────────────────────────────────────────────
with tab_files:
    st.markdown("## 📂 Files & Preprocessing")
    st.caption(
        "Upload your linescan files and configure preprocessing and spectral range. "
        "These settings apply to both PLS and MCR analysis. "
        "Then go to the **PLS regression** or **MCR decomposition** tab to run the analysis."
    )
    _fc1, _fc2 = st.columns([1, 1])

    with _fc1:
        st.subheader("Linescan files")
        linescan_files = st.file_uploader(
            "Linescan .txt files",
            type=["txt"],
            accept_multiple_files=True,
            help="Raw linescan files from the Raman microscope (one or more).",
        )
        if linescan_files:
            st.caption(f"{len(linescan_files)} file(s) uploaded: "
                       + ", ".join(f.name for f in linescan_files))

        st.subheader("Preprocessing")
        baseline = st.selectbox(
            "Baseline correction",
            ["rubberband", "rolling_ball", "als", "arpls", "airpls", "iasls", "drpls", "imodpoly", "modpoly", "poly", "endpoint", "linear", "none"],
            format_func=lambda x: {
                "rubberband":  "Rubberband (convex hull)",
                "rolling_ball":"Rolling ball",
                "als":         "ALS (Asymmetric Least Squares)",
                "arpls":       "ARPLS (Asymmetrically Reweighted PLS)",
                "airpls":      "AIRPLS (Adaptive Iterative Reweighted PLS)",
                "iasls":       "IASLS (Improved ALS)",
                "drpls":       "DRPLS (Doubly Reweighted PLS)",
                "imodpoly":    "IModPoly (Improved Modified Polynomial)",
                "modpoly":     "ModPoly (Modified Polynomial)",
                "poly":        "Poly (Polynomial)",
                "endpoint":    "Endpoint (anchors ends to zero)",
                "linear":      "Linear (fit to spectral edges)",
                "none":        "None",
            }[x],
            help=(
                "Removes fluorescence background. "
                "**Rubberband** (default) — fast convex hull, robust for most Raman data. "
                "**ARPLS / AIRPLS / IASLS / DRPLS** — penalised least-squares variants from RamanSPy. "
                "**IModPoly / ModPoly / Poly** — polynomial fits. "
                "**Rolling ball** — morphological, tunable radius. "
                "**ALS** — classic asymmetric least squares. "
                "**Endpoint** — straight line between first and last point. "
                "**Linear** — straight line fitted to outermost 5 % of points."
            ),
        )
        ball_radius, als_lam, als_p = 50, 1e5, 0.01
        rs_lam, rs_poly_order = 1e5, 2
        if baseline == "rolling_ball":
            ball_radius = st.slider("Ball radius", 10, 300, 50)
        elif baseline == "als":
            als_lam = st.number_input("λ (smoothness)", value=1e5, min_value=1e2, max_value=1e8, format="%.0e")
            als_p   = st.number_input("p (asymmetry)",  value=0.01, min_value=0.001, max_value=0.5, format="%.3f")
        elif baseline in ("arpls", "airpls", "iasls", "drpls"):
            rs_lam = st.number_input("λ (smoothness)", value=1e5, min_value=1e2, max_value=1e9, format="%.0e",
                                     key="rs_lam")
        elif baseline in ("imodpoly", "modpoly", "poly"):
            rs_poly_order = st.slider("Polynomial order", 1, 8, 2, key="rs_poly")

        smooth = st.selectbox(
            "Smoothing",
            ["none", "savgol", "gaussian", "fft_lowpass"],
            format_func=lambda x: {
                "none":        "None",
                "savgol":      "Savitzky-Golay",
                "gaussian":    "Gaussian",
                "fft_lowpass": "FFT low-pass filter",
            }[x],
            help=(
                "Optional smoothing applied after baseline correction. "
                "**Savitzky-Golay** — polynomial smoothing, preserves peak shapes well. "
                "**Gaussian** — convolves with a Gaussian kernel. "
                "**FFT low-pass** — removes high-frequency noise by zeroing Fourier components above a cutoff fraction."
            ),
        )
        sg_window, sg_poly, gaussian_sigma, fft_cutoff = 11, 3, 1, 0.1
        if smooth == "savgol":
            sg_window = st.slider("Window length (odd)", 5, 51, 11, step=2, key="sg_win_main")
            sg_poly   = st.slider("Polynomial order",    1,  9,  3,        key="sg_poly_main")
        elif smooth == "gaussian":
            gaussian_sigma = st.slider("σ (sigma)", 0.5, 5.0, 1.0, step=0.5, key="gauss_sigma")
        elif smooth == "fft_lowpass":
            fft_cutoff = st.slider("Cutoff fraction", 0.01, 0.5, 0.1, step=0.01, key="fft_cut",
                                   help="Fraction of Fourier components to keep (0.1 = keep lowest 10%).")

        _norm_opts = ["snv", "area", "vector", "minmax", "none"]
        _norm_fmt  = lambda x: {
            "snv":    "SNV",
            "area":   "Area (unit area)",
            "vector": "Vector (unit norm)",
            "minmax": "Min-max [0, 1]",
            "none":   "None",
        }[x]
        normalize_pls = st.selectbox(
            "PLS regression normalization",
            _norm_opts,
            index=0,
            format_func=_norm_fmt,
            help=(
                "Normalization applied to spectra before PLS modelling. "
                "**SNV** (default) — centres and scales to zero mean and unit variance; "
                "the standard choice for PLS. "
                "**Area** — divides by spectral integral; removes intensity differences while "
                "preserving relative peak ratios. "
                "**Vector** — divides by Euclidean norm; ensures all spectra have equal L2 norm. "
                "**Min-max** — not recommended for PLS as a single spike can compress all features."
            ),
        )
        normalize_mcr = st.selectbox(
            "MCR-ALS normalization",
            _norm_opts,
            index=1,
            format_func=_norm_fmt,
            help=(
                "Normalization applied to spectra before MCR-ALS decomposition. "
                "**Area** (default) — recommended for MCR as it preserves non-negativity and "
                "ensures intensity differences reflect concentration rather than instrument artifacts. "
                "**SNV** can cause MCR to converge in 1 iteration because it centres spectra around "
                "zero, conflicting with non-negativity constraints. "
                "**Vector** — normalises to unit Euclidean norm."
            ),
        )
        normalize = normalize_mcr  # used by salt and further-analysis paths
        salt_normalize = st.selectbox(
            "Salt PLS regression normalization",
            ["none", "minmax", "snv"],
            format_func=_norm_fmt,
            help="Normalization applied only to salt standards and linescan salt preprocessing.",
        )
        spike_remove = st.toggle("Spike removal", value=True)

    with _fc2:
        st.subheader("Spectral Range")
        c1, c2 = st.columns(2)
        wn_min = c1.number_input("Protein min (cm⁻¹)", value=700,  step=50, key="wn_min")
        wn_max = c2.number_input("Protein max (cm⁻¹)", value=3900, step=50, key="wn_max")

        use_cut = st.toggle("Exclude gap region", value=True, key="use_cut",
                            help="E.g. remove the 1850–2750 cm⁻¹ silent region.")
        wn_cut_min = wn_cut_max = None
        if use_cut:
            c3, c4 = st.columns(2)
            wn_cut_min = c3.number_input("Gap from", value=1850, step=50, key="wn_cut_min")
            wn_cut_max = c4.number_input("Gap to",   value=2750, step=50, key="wn_cut_max")

        c5, c6 = st.columns(2)
        salt_wn_min = c5.number_input("Salt min (cm⁻¹)", value=900,  step=10, key="salt_wn_min",
                                       help="Lower bound of the wavenumber region used for salt PLS.")
        salt_wn_max = c6.number_input("Salt max (cm⁻¹)", value=1000, step=10, key="salt_wn_max",
                                       help="Upper bound of the salt region.")

        st.subheader("Concentration unit")
        unit = st.selectbox("Protein concentration unit", ["mg/mL", "mM"])
        protein_mw = 5808.0
        if unit == "mM":
            protein_mw = float(st.number_input("Protein MW (g/mol)", value=5808, min_value=100))
        conv = 1000.0 / protein_mw if unit == "mM" else 1.0


# ─────────────────────────────────────────────────────────────────────────────
# PLS regression tab — configuration at top
# ─────────────────────────────────────────────────────────────────────────────
with tab_pls:
    st.markdown("## 📊 PLS Calibration")
    st.caption(
        "Upload standard CSV files and set cross-validation parameters, then click "
        "**▶ Build PLS Model**. Calibration results will appear below."
    )
    _pc1, _pc2 = st.columns([1, 1])

    with _pc1:
        st.subheader("Standard spectra")
        protein_std_src = _std_picker(
            "Protein standard CSV",
            "Col 0: concentration (mg/mL), remaining cols: spectra.",
            "prot", "Protein",
        )
        peg_std_src = _std_picker(
            "Molecular crowder standard CSV",
            "Col 0: concentration (wt%), remaining cols: spectra. "
            "When provided, a dual protein+molecular crowder PLS model is trained.",
            "peg", "PEG", optional=True,
        )
        salt_std_src = _std_picker(
            "Salt standard CSV",
            "Col 0: concentration, remaining cols: spectra.",
            "salt", "Salt", optional=True,
        )

    with _pc2:
        st.subheader("Cross-validation")
        _cv1, _cv2 = st.columns(2)
        cv_folds = _cv1.number_input(
            "CV folds", min_value=2, max_value=20, value=5, step=1,
            help="Number of folds for k-fold cross-validation during PLS component selection. "
                 "Default: 5.",
        )
        max_pls_components = _cv2.number_input(
            "Max LVs", min_value=1, max_value=50, value=20, step=1,
            help="Maximum number of latent variables evaluated during CV. "
                 "The optimum is selected automatically by minimising CV RMSE.",
        )

    build_btn = st.button(
        "▶ Build PLS Model",
        use_container_width=True,
        type="primary",
        disabled=not linescan_files,
        help="Upload linescan files first (Files tab) to enable.",
        on_click=_goto_pls,
    )
    st.divider()
    st.markdown("### Results")


# ─────────────────────────────────────────────────────────────────────────────
# MCR decomposition tab — configuration at top
# ─────────────────────────────────────────────────────────────────────────────
with tab_mcr:
    st.markdown("## 🔬 MCR Decomposition")
    st.caption(
        "Configure MCR-ALS initialisation and parameters, then click "
        "**▶ Run MCR Analysis**. Concentration profiles and recovered spectra will appear below."
    )
    _mc1, _mc2 = st.columns([1, 1])

    with _mc1:
        st.subheader("Initialisation")
        mcr_init_mode = st.radio(
            "MCR initialisation",
            ["Reference CSV", "PCA from linescans"],
            key="mcr_init_mode",
            help=(
                "**Reference CSV**: provide known pure-component spectra to initialise MCR-ALS. "
                "Recommended when reference spectra are available. "
                "**PCA from linescans**: top principal components of each preprocessed linescan "
                "are used as the initial ST estimate."
            ),
        )
        mcr_ref_src  = None
        mcr_comp_ids = None
        if mcr_init_mode == "Reference CSV":
            mcr_ref_src = _std_picker(
                "MCR reference spectra CSV",
                "Col 0: component label, remaining cols: reference spectra.",
                "mcr", None, optional=True,
            )
            if mcr_ref_src:
                try:
                    _df_mcr  = _read_csv_src(mcr_ref_src)
                    _all_ids = _df_mcr.iloc[:, 0].tolist()
                    _sel = st.multiselect(
                        "Select MCR components",
                        options=_all_ids,
                        default=_all_ids,
                        key="mcr_comp_sel",
                        help="Choose which reference spectra rows to use.",
                    )
                    mcr_comp_ids = _sel if _sel else _all_ids
                except Exception:
                    pass
        else:
            st.caption(
                "PCA will be computed from each linescan at run time. "
                "Set the number of components to match the expected chemical contributors."
            )
        n_components = st.slider(
            "MCR components", 1, 10, 3,
            help="Number of pure-component spectra MCR-ALS will resolve.",
        )
        if mcr_init_mode == "Reference CSV" and mcr_ref_src:
            try:
                _n_ref_rows = len(_df_mcr) if mcr_comp_ids is None else len(mcr_comp_ids)
            except Exception:
                _n_ref_rows = 0
            if _n_ref_rows and n_components > _n_ref_rows:
                st.warning(
                    f"⚠ {n_components} components requested but only {_n_ref_rows} reference "
                    f"spectrum/spectra selected. MCR-ALS will use all {_n_ref_rows} references."
                )

    with _mc2:
        st.subheader("MCR-ALS parameters")
        mcr_max_iter = st.number_input(
            "Max iterations", min_value=10, max_value=20000, value=2000, step=100,
            help="Maximum number of ALS iterations. Default: 2000.",
        )
        mcr_tol = st.number_input(
            "Convergence tolerance", min_value=1e-6, max_value=1.0, value=1e-2,
            format="%.2e",
            help="Relative change in fit residual below which MCR-ALS is considered converged. Default: 0.01.",
        )
        _mr1, _mr2 = st.columns(2)
        mcr_c_regr = _mr1.selectbox(
            "C regression", ["OLS", "NNLS"],
            help="Regression method for updating the concentration matrix C.",
        )
        mcr_st_regr = _mr2.selectbox(
            "ST regression", ["NNLS", "OLS"],
            help="Regression method for updating the spectra matrix ST. "
                 "NNLS enforces non-negative spectral values.",
        )

    run_btn = st.button(
        "▶ Run MCR Analysis",
        use_container_width=True,
        type="primary",
        disabled=not linescan_files,
        help="Upload linescan files first (Files tab) to enable.",
        on_click=_goto_mcr,
    )
    st.divider()
    st.markdown("### Results")

# Collect settings into a dict used by analysis.py
settings = dict(
    baseline=baseline,
    ball_radius=ball_radius,
    als_lam=als_lam,
    als_p=als_p,
    rs_lam=rs_lam,
    rs_poly_order=rs_poly_order,
    smooth=smooth,
    sg_window=sg_window,
    sg_poly=sg_poly,
    gaussian_sigma=gaussian_sigma,
    fft_cutoff=fft_cutoff,
    normalize=normalize_mcr,
    normalize_pls=normalize_pls,
    normalize_mcr=normalize_mcr,
    salt_normalize=salt_normalize,
    spike_remove=spike_remove,
    wn_min=wn_min,
    wn_max=wn_max,
    use_cut=use_cut,
    wn_cut_min=wn_cut_min,
    wn_cut_max=wn_cut_max,
    salt_wn_min=salt_wn_min,
    salt_wn_max=salt_wn_max,
)
# Settings override for PLS standard preprocessing
pls_settings = dict(settings, normalize=normalize_pls)


# ─────────────────────────────────────────────────────────────────────────────
# Build Models
# ─────────────────────────────────────────────────────────────────────────────
if build_btn:
    if not linescan_files:
        st.toast("Upload at least one linescan file to derive the wavenumber axis.", icon="❌")
    else:
        try:
            with st.spinner("Deriving wavenumber axis from first linescan…"):
                first = linescan_files[0]
                wn_ref, _, _ = an.load_linescan_bytes(first.read(), first.name)
                first.seek(0)

            # ── Protein PLS (optional) ───────────────────────────────────────
            pls_protein = None
            if protein_std_src:
                with st.spinner("Loading & preprocessing protein standards…"):
                    df_prot = _read_csv_src(protein_std_src)
                    y_prot = df_prot.iloc[:, 0].to_numpy(dtype=float) * conv
                    X_prot_raw = df_prot.iloc[:, 1:].to_numpy()
                    X_prot_proc, wn_prot, _ = an.preprocess_matrix(X_prot_raw, wn_ref, pls_settings)

                # ── Dual PLS (protein + PEG) or single-output protein PLS ───
                if peg_std_src:
                    with st.spinner("Loading & preprocessing molecular crowder standards…"):
                        df_peg = _read_csv_src(peg_std_src)
                        y_peg = df_peg.iloc[:, 0].to_numpy(dtype=float)
                        X_peg_raw = df_peg.iloc[:, 1:].to_numpy()
                        X_peg_proc, _, _ = an.preprocess_matrix(X_peg_raw, wn_ref, pls_settings)

                    with st.spinner("Training dual protein+molecular crowder PLS model…"):
                        pls_protein = an.build_dual_pls_model(
                            X_prot_proc, y_prot, X_peg_proc, y_peg,
                            max_components=int(max_pls_components), cv_folds=int(cv_folds),
                        )
                        pls_protein["wn"] = wn_prot
                        pls_protein["X_prot_proc"] = X_prot_proc
                        pls_protein["X_peg_proc"]  = X_peg_proc
                        pls_protein["y_prot_raw"]  = y_prot
                        pls_protein["y_peg_raw"]   = y_peg
                else:
                    with st.spinner("Training protein PLS model…"):
                        pls_protein = an.build_pls_model(
                            X_prot_proc, y_prot,
                            max_components=int(max_pls_components), cv_folds=int(cv_folds),
                        )
                        pls_protein["wn"] = wn_prot
                        pls_protein["X_train_proc"] = X_prot_proc
                        pls_protein["y_raw"] = y_prot

            # ── Salt PLS (optional, fingerprint region) ─────────────────────
            pls_salt = None
            if salt_std_src:
                with st.spinner("Loading & preprocessing salt standards…"):
                    df_salt = _read_csv_src(salt_std_src)
                    y_salt = df_salt.iloc[:, 0].to_numpy(dtype=float)
                    X_salt_raw = df_salt.iloc[:, 1:].to_numpy()
                    X_salt_proc, wn_salt, _ = an.preprocess_matrix(
                        X_salt_raw, wn_ref, settings, is_salt=True
                    )
                with st.spinner("Training salt PLS model…"):
                    pls_salt = an.build_pls_model(
                        X_salt_proc, y_salt,
                        max_components=int(max_pls_components), cv_folds=int(cv_folds),
                    )
                    pls_salt["wn"] = wn_salt
                    pls_salt["X_train_proc"] = X_salt_proc
                    pls_salt["y_raw"] = y_salt

            # ── MCR references (optional) ───────────────────────────────────
            ST_init = None
            comp_labels = []
            if mcr_ref_src:
                with st.spinner("Preprocessing MCR references…"):
                    df_refs = _read_csv_src(mcr_ref_src)
                    all_ids = df_refs.iloc[:, 0].tolist()
                    selected_ids = mcr_comp_ids if mcr_comp_ids else all_ids
                    row_mask = [i for i, cid in enumerate(all_ids) if cid in selected_ids]
                    comp_raw = df_refs.iloc[row_mask, 1:].to_numpy()
                    comp_labels = [str(all_ids[i]) for i in row_mask]
                    ST_init, _, _ = an.preprocess_matrix(comp_raw, wn_ref, settings)

            st.session_state["models"] = dict(
                pls_protein=pls_protein,
                pls_salt=pls_salt,
                ST_init=ST_init,
                wn_ref=wn_ref,
                comp_labels=comp_labels,
            )
            st.session_state["settings"]      = settings
            st.session_state["n_components"]  = n_components
            st.session_state["unit"]          = unit
            st.session_state["mcr_params"]    = dict(
                max_iter=int(mcr_max_iter),
                tol_increase=float(mcr_tol),
                c_regr=mcr_c_regr,
                st_regr=mcr_st_regr,
            )

            msg_parts = []
            if pls_protein is not None:
                dual = pls_protein.get("dual", False)
                if dual:
                    msg_parts.append(
                        f"Dual protein+molecular crowder PLS: {pls_protein['n_components']} LVs, "
                        f"CV RMSE — Protein: {pls_protein['cv_rmse_protein']:.3f} {unit}, "
                        f"Molecular crowder: {pls_protein['cv_rmse_peg']:.3f} wt%"
                    )
                else:
                    msg_parts.append(
                        f"Protein PLS: {pls_protein['n_components']} LVs, "
                        f"CV RMSE = {pls_protein['cv_rmse']:.3f} {unit}"
                    )
            if pls_salt is not None:
                msg_parts.append(
                    f"Salt PLS: {pls_salt['n_components']} LVs, "
                    f"CV RMSE = {pls_salt['cv_rmse']:.3f}"
                )
            if ST_init is not None:
                msg_parts.append(f"MCR refs: {ST_init.shape[0]} component(s) preprocessed")
            if not msg_parts:
                msg_parts.append("Wavenumber axis loaded (no standards provided)")
            st.toast("\n".join(msg_parts), icon="✅")

            # Immediately apply PLS to all uploaded linescans (MCR skipped)
            if pls_protein is not None or pls_salt is not None:
                _auto_results = {}
                _auto_prog = st.progress(0, text="Applying PLS to linescans…")
                _auto_sm = "xy"
                for _ai, _af in enumerate(linescan_files):
                    _auto_prog.progress(_ai / len(linescan_files), text=f"Processing {_af.name}…")
                    try:
                        _awn, _aX, _apos = an.load_linescan_bytes(_af.read(), _af.name)
                        _af.seek(0)
                        _auto_sm = an.detect_scan_mode(_apos)
                        _auto_results[_af.name] = an.process_linescan(
                            _awn, _aX, _apos,
                            scan_mode=_auto_sm,
                            pls_protein_info=pls_protein,
                            pls_salt_info=pls_salt,
                            ST_init=None,
                            n_components=n_components,
                            settings=settings,
                            pls_settings=pls_settings,
                        )
                    except Exception as _ae:
                        st.warning(f"Could not process {_af.name}: {_ae}")
                _auto_prog.progress(1.0, text="Done!")
                st.session_state["results"]   = _auto_results
                st.session_state["scan_mode"] = _auto_sm
            else:
                st.session_state.pop("results", None)

        except Exception as exc:
            st.toast(f"Error building models: {exc}", icon="❌")
            raise


# ─────────────────────────────────────────────────────────────────────────────
# Run Analysis
# ─────────────────────────────────────────────────────────────────────────────
if run_btn:
    if not linescan_files:
        st.toast("Upload linescan files first.", icon="❌")
    else:
        models        = st.session_state.get("models", {})
        s             = st.session_state.get("settings", settings)
        n_comp        = n_components
        mcr_params    = st.session_state.get("mcr_params", dict(
            max_iter=int(mcr_max_iter),
            tol_increase=float(mcr_tol),
            c_regr=mcr_c_regr,
            st_regr=mcr_st_regr,
        ))
        # Use current sidebar value if not stored in session_state (i.e. Build not clicked)
        _mcr_init_mode = st.session_state.get("mcr_init_mode", mcr_init_mode)
        results       = {}

        prog = st.progress(0, text="Processing linescans…")
        for i, f in enumerate(linescan_files):
            prog.progress(i / len(linescan_files), text=f"Processing {f.name}…")
            try:
                wn, X, positions = an.load_linescan_bytes(f.read(), f.name)
                f.seek(0)
                sm = an.detect_scan_mode(positions)

                # Determine ST_init for this file
                _pca_var_exp = None
                _pca_loadings_stored = None
                _pca_wn_stored = None
                if _mcr_init_mode == "PCA from linescans":
                    from sklearn.decomposition import PCA as _PCA
                    _X_pre, _wn_pre, _ = an.preprocess_matrix(X, wn, s)
                    _n_pca = min(n_comp, _X_pre.shape[0], _X_pre.shape[1])
                    _pca = _PCA(n_components=_n_pca)
                    _pca.fit(_X_pre)
                    # Ensure each loading is non-negative: flip sign if majority is negative
                    _loadings = _pca.components_.copy()
                    for _k in range(_loadings.shape[0]):
                        if _loadings[_k].sum() < 0:
                            _loadings[_k] *= -1
                        _loadings[_k] = np.clip(_loadings[_k], 0, None)
                    _st_init = _loadings
                    _pca_var_exp = _pca.explained_variance_ratio_ * 100
                    _pca_loadings_stored = _pca.components_
                    _pca_wn_stored = _wn_pre
                else:
                    # Re-process MCR refs directly from the selected source.
                    # wn_ref from models is preferred (set by Build); fall back to
                    # the current linescan's wn so MCR works even if Build was skipped.
                    _st_init = models.get("ST_init")
                    if mcr_ref_src:
                        try:
                            _df_mcr_run = _read_csv_src(mcr_ref_src)
                            _all_ids_run = _df_mcr_run.iloc[:, 0].tolist()
                            _sel_ids_run = mcr_comp_ids if mcr_comp_ids else _all_ids_run
                            _rmask_run = [j for j, cid in enumerate(_all_ids_run) if cid in _sel_ids_run]
                            _comp_raw_run = _df_mcr_run.iloc[_rmask_run, 1:].to_numpy()
                            _labels_run = [str(_all_ids_run[j]) for j in _rmask_run]
                            _wn_for_refs = models.get("wn_ref", wn)
                            _st_init, _, _ = an.preprocess_matrix(_comp_raw_run, _wn_for_refs, s)
                            if "models" not in st.session_state:
                                st.session_state["models"] = {}
                            st.session_state["models"]["ST_init"] = _st_init
                            st.session_state["models"]["comp_labels"] = _labels_run
                        except Exception as _re:
                            st.warning(f"Could not re-process MCR references: {_re}")

                _res = an.process_linescan(
                    wn, X, positions,
                    scan_mode=sm,
                    pls_protein_info=models.get("pls_protein"),
                    pls_salt_info=models.get("pls_salt"),
                    ST_init=_st_init,
                    n_components=n_comp,
                    settings=s,
                    mcr_params=mcr_params,
                    pls_settings=pls_settings,
                )
                if _pca_var_exp is not None:
                    _res["pca_var_exp"]  = _pca_var_exp
                    _res["pca_loadings"] = _pca_loadings_stored
                    _res["pca_wn"]       = _pca_wn_stored
                results[f.name] = _res
            except Exception as exc:
                st.warning(f"Could not process {f.name}: {exc}")
        prog.progress(1.0, text="Done!")
        st.session_state["results"]   = results
        st.session_state["scan_mode"] = sm
        # When PCA was used for MCR initialisation, overwrite comp_labels with
        # generic names so the results tab doesn't show CSV-derived labels.
        if _mcr_init_mode == "PCA from linescans":
            _pca_labels = [f"comp_{k + 1}" for k in range(n_comp)]
            if "models" not in st.session_state:
                st.session_state["models"] = {}
            st.session_state["models"]["comp_labels"] = _pca_labels



# ── Preview ──────────────────────────────────────────────────────────────────
with tab_preview:
    st.subheader("Uploaded data preview")
    if not linescan_files:
        st.info("Upload linescan files using the controls above.")
    else:
        st.markdown(f"**{len(linescan_files)} linescan file(s)** ready")
        if st.button("Preview first linescan (preprocessed)"):
            try:
                first = linescan_files[0]
                wn, X, _ = an.load_linescan_bytes(first.read(), first.name)
                first.seek(0)
                n_show = min(X.shape[0], 60)

                def _prev_plot(X_p, wn_p, title, color):
                    fig = go.Figure()
                    for i in range(min(X_p.shape[0], n_show)):
                        fig.add_trace(go.Scatter(
                            x=wn_p, y=X_p[i], mode="lines",
                            line=dict(color=color, width=0.8),
                            opacity=0.4, showlegend=False,
                        ))
                    fig.update_layout(
                        title=title,
                        xaxis_title="Wavenumber (cm⁻¹)",
                        yaxis_title="Intensity (a.u.)",
                        height=320, margin=dict(t=40, b=40),
                    )
                    return fig

                with st.spinner("Preprocessing…"):
                    X_mcr,  wn_mcr,  _ = an.preprocess_matrix(X, wn, settings)
                    X_pls,  wn_pls,  _ = an.preprocess_matrix(X, wn, pls_settings)
                    X_salt, wn_salt, _ = an.preprocess_matrix(X, wn,
                        dict(settings, normalize=settings.get("salt_normalize", "none")),
                        is_salt=True)

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.plotly_chart(_prev_plot(X_pls,  wn_pls,
                        f"PLS regression normalization ({normalize_pls})", COLORS[0]),
                        use_container_width=True)
                    st.caption(f"{X_pls.shape[0]} spectra, {wn_pls[0]:.0f}–{wn_pls[-1]:.0f} cm⁻¹")
                with c2:
                    st.plotly_chart(_prev_plot(X_mcr,  wn_mcr,
                        f"MCR-ALS normalization ({normalize_mcr})", COLORS[1]),
                        use_container_width=True)
                    st.caption(f"{X_mcr.shape[0]} spectra, {wn_mcr[0]:.0f}–{wn_mcr[-1]:.0f} cm⁻¹")
                with c3:
                    st.plotly_chart(_prev_plot(X_salt, wn_salt,
                        f"Salt PLS normalization ({salt_normalize})", COLORS[3]),
                        use_container_width=True)
                    st.caption(f"{X_salt.shape[0]} spectra, {wn_salt[0]:.0f}–{wn_salt[-1]:.0f} cm⁻¹")

            except Exception as exc:
                st.error(f"Preview failed: {exc}")


# ── Calibration ───────────────────────────────────────────────────────────────
with tab_calib:
    if "models" not in st.session_state:
        st.info("Build a PLS model first — upload standard CSVs above and click ▶ Build PLS Model.")
    else:
        models = st.session_state["models"]
        unit   = st.session_state.get("unit", "mg/mL")
        pls_p  = models.get("pls_protein")
        pls_s  = models.get("pls_salt")
        dual   = pls_p.get("dual", False) if pls_p else False

        if pls_p is None:
            st.info("No protein PLS model was built. Upload a protein standard CSV above and click ▶ Build PLS Model.")


        if pls_p is not None and dual:
            # ── Dual protein + molecular crowder calibration ────────────────
            st.subheader("Dual Protein + Molecular Crowder PLS regression")
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Components", pls_p["n_components"])
            m2.metric(f"Protein CV RMSE", f"{pls_p['cv_rmse_protein']:.4f} {unit}")
            m3.metric("Protein Train R²", f"{pls_p['r2_protein_train']:.4f}")
            m4.metric("Crowder CV RMSE",  f"{pls_p['cv_rmse_peg']:.4f} wt%")
            m5.metric("Crowder Train R²", f"{pls_p['r2_peg_train']:.4f}")

            fig_d = make_subplots(
                rows=2, cols=2,
                subplot_titles=[
                    f"a)  Protein calibration (actual vs predicted, {unit})",
                    "b)  Molecular crowder calibration (actual vs predicted, wt%)",
                    "c)  CV vs Training RMSE",
                    "d)  PLS loadings (all PCs)",
                ],
            )
            wn_p = pls_p["wn"]
            y_prot_tr = pls_p["y_protein_train"]
            y_prot_pr = pls_p["y_pred_protein_train"]
            y_peg_tr  = pls_p["y_peg_train"]
            y_peg_pr  = pls_p["y_pred_peg_train"]

            # Protein scatter
            fig_d.add_trace(go.Scatter(
                x=y_prot_tr, y=y_prot_pr, mode="markers",
                marker=dict(color=COLORS[0], size=7, line=dict(color="black", width=0.5)),
                name="Protein", showlegend=True,
            ), row=1, col=1)
            fig_d.add_trace(go.Scatter(
                x=[y_prot_tr.min(), y_prot_tr.max()],
                y=[y_prot_tr.min(), y_prot_tr.max()],
                mode="lines", line=dict(dash="dash", color="black", width=1), showlegend=False,
            ), row=1, col=1)

            # Molecular crowder scatter
            fig_d.add_trace(go.Scatter(
                x=y_peg_tr, y=y_peg_pr, mode="markers",
                marker=dict(color=COLORS[2], size=7, line=dict(color="black", width=0.5)),
                name="Molecular crowder", showlegend=True,
            ), row=1, col=2)
            fig_d.add_trace(go.Scatter(
                x=[y_peg_tr.min(), y_peg_tr.max()],
                y=[y_peg_tr.min(), y_peg_tr.max()],
                mode="lines", line=dict(dash="dash", color="black", width=1), showlegend=False,
            ), row=1, col=2)

            # CV / Train RMSE
            n_cv = len(pls_p["rmse_cv_all"])
            comps = list(range(1, n_cv + 1))
            fig_d.add_trace(go.Scatter(
                x=comps, y=pls_p["rmse_cv_all"], mode="lines+markers",
                name="CV RMSE", line=dict(color="red"),
            ), row=2, col=1)
            fig_d.add_trace(go.Scatter(
                x=comps, y=pls_p["rmse_train_all"], mode="lines+markers",
                name="Train RMSE", line=dict(color="steelblue"),
            ), row=2, col=1)
            fig_d.add_vline(
                x=pls_p["n_components"], line_dash="dash", line_color="black",
                annotation_text=f"opt={pls_p['n_components']}", row=2, col=1,
            )
            if n_cv > pls_p["n_components"]:
                fig_d.add_vrect(
                    x0=pls_p["n_components"] + 0.5, x1=n_cv + 0.5,
                    fillcolor="rgba(255,0,0,0.12)", line_width=0,
                    annotation_text="Overfitting", annotation_position="top right",
                    annotation=dict(font_size=11, font_color="red", y=0.5, yanchor="middle"),
                    row=2, col=1,
                )

            # Loadings — all components
            valid = pls_p["valid_features"]
            if wn_p is not None:
                wn_valid = wn_p[valid]
                for i in range(pls_p["n_components"]):
                    fig_d.add_trace(go.Scatter(
                        x=wn_valid, y=pls_p["model"].x_loadings_[:, i],
                        mode="lines", name=f"PC{i + 1}",
                    ), row=2, col=2)

            fig_d.update_xaxes(title_text=f"Actual ({unit})",   row=1, col=1)
            fig_d.update_xaxes(title_text="Actual (wt%)",       row=1, col=2)
            fig_d.update_xaxes(title_text="N components",       row=2, col=1)
            fig_d.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=2, col=2)
            fig_d.update_yaxes(title_text=f"Predicted ({unit})", row=1, col=1)
            fig_d.update_yaxes(title_text="Predicted (wt%)",    row=1, col=2)
            fig_d.update_yaxes(title_text="RMSE",               row=2, col=1)
            fig_d.update_yaxes(title_text="Loading",            row=2, col=2)
            fig_d.update_layout(height=580, legend=dict(orientation="h", y=-0.12))
            st.plotly_chart(fig_d, use_container_width=True)
            st.caption(
                f"PLS2 calibration using {pls_p['n_components']} latent variables to simultaneously predict "
                f"protein and molecular crowder concentration. "
                f"**a)** Protein: predicted vs actual ({unit}), train R² = {pls_p['r2_protein_train']:.4f}. "
                f"**b)** Molecular crowder: predicted vs actual (wt%), train R² = {pls_p['r2_peg_train']:.4f}. "
                f"**c)** Leave-one-out CV RMSE (red) and training RMSE (blue) as a function of the number of latent variables; "
                f"the dashed line marks the selected optimum (opt = {pls_p['n_components']}). "
                f"Protein CV RMSE = {pls_p['cv_rmse_protein']:.4f} {unit}; Molecular crowder CV RMSE = {pls_p['cv_rmse_peg']:.4f} wt%. "
                f"**d)** X-loadings for each latent variable, highlighting the spectral features driving the model."
            )

            # PLS coefficients
            if wn_p is not None:
                with st.expander("PLS regression coefficients"):
                    wn_valid = wn_p[pls_p["valid_features"]]
                    coef = _pls_coef(pls_p["model"], len(wn_valid))
                    fig_coef = go.Figure()
                    fig_coef.add_trace(go.Scatter(
                        x=wn_valid, y=coef[:, 0], mode="lines",
                        name="Protein", line=dict(color=COLORS[0], width=1.5),
                    ))
                    fig_coef.add_trace(go.Scatter(
                        x=wn_valid, y=coef[:, 1], mode="lines",
                        name="Molecular crowder", line=dict(color=COLORS[2], width=1.5),
                    ))
                    fig_coef.add_hline(y=0, line_dash="dot", line_color="grey")
                    fig_coef.update_layout(
                        xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Coefficient",
                        height=300, legend=dict(orientation="h"),
                    )
                    st.plotly_chart(fig_coef, use_container_width=True)
                    st.caption(
                        f"PLS2 regression coefficient vectors for protein ({unit}, blue) and molecular crowder (wt%, green) "
                        f"over {wn_valid[0]:.0f}–{wn_valid[-1]:.0f} cm⁻¹. "
                        f"Positive coefficients indicate spectral features positively correlated with concentration; "
                        f"negative coefficients indicate anticorrelated features. "
                        f"Peaks in the protein vector reflect protein-specific Raman bands."
                    )

            # Preprocessed training spectra — separate plots
            with st.expander("Training spectra"):
                X_tr_all = pls_p["X_train_proc"]
                wn_tr    = pls_p["wn"]
                n_prot   = pls_p["X_prot_proc"].shape[0]
                _n_peg   = X_tr_all.shape[0] - n_prot

                st.markdown("**Protein standards**")
                fig_sp_prot = go.Figure()
                for i in range(n_prot):
                    fig_sp_prot.add_trace(go.Scatter(
                        x=wn_tr, y=X_tr_all[i], mode="lines",
                        line=dict(color=COLORS[0], width=0.8), opacity=0.5,
                        showlegend=False,
                    ))
                fig_sp_prot.update_layout(
                    xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Norm. intensity", height=250,
                )
                st.plotly_chart(fig_sp_prot, use_container_width=True)
                st.caption(
                    f"{n_prot} preprocessed protein standard spectra covering {wn_tr[0]:.0f}–{wn_tr[-1]:.0f} cm⁻¹."
                )

                st.markdown("**Molecular crowder standards**")
                fig_sp_peg = go.Figure()
                for i in range(n_prot, X_tr_all.shape[0]):
                    fig_sp_peg.add_trace(go.Scatter(
                        x=wn_tr, y=X_tr_all[i], mode="lines",
                        line=dict(color=COLORS[2], width=0.8), opacity=0.5,
                        showlegend=False,
                    ))
                fig_sp_peg.update_layout(
                    xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Norm. intensity", height=250,
                )
                st.plotly_chart(fig_sp_peg, use_container_width=True)
                st.caption(
                    f"{_n_peg} preprocessed molecular crowder standard spectra covering {wn_tr[0]:.0f}–{wn_tr[-1]:.0f} cm⁻¹."
                )

                if pls_s is not None and "X_train_proc" in pls_s:
                    st.markdown("**Salt standards**")
                    fig_sp_salt = go.Figure()
                    for i in range(pls_s["X_train_proc"].shape[0]):
                        fig_sp_salt.add_trace(go.Scatter(
                            x=pls_s["wn"], y=pls_s["X_train_proc"][i], mode="lines",
                            line=dict(color=COLORS[3], width=0.8), opacity=0.5,
                            showlegend=False,
                        ))
                    fig_sp_salt.update_layout(
                        xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Norm. intensity", height=250,
                    )
                    st.plotly_chart(fig_sp_salt, use_container_width=True)
                    st.caption(
                        f"{pls_s['X_train_proc'].shape[0]} preprocessed salt standard spectra "
                        f"covering {pls_s['wn'][0]:.0f}–{pls_s['wn'][-1]:.0f} cm⁻¹."
                    )

        elif pls_p is not None:
            # ── Single-output protein PLS calibration ─────────
            st.subheader("Protein PLS regression")
            m1, m2, m3 = st.columns(3)
            m1.metric("Optimal components", pls_p["n_components"])
            m2.metric("CV RMSE",  f"{pls_p['cv_rmse']:.4f} {unit}")
            m3.metric("Train R²", f"{pls_p['r2_train']:.4f}")

            fig_p = make_subplots(
                rows=2, cols=2,
                subplot_titles=[
                    "a)  Preprocessed standards",
                    f"b)  Calibration (actual vs predicted, {unit})",
                    "c)  CV vs Training RMSE",
                    "d)  PLS loadings (all PCs)",
                ],
            )
            wn_p   = pls_p["wn"]
            X_tr   = pls_p["X_train_proc"]
            y_tr   = pls_p["y_raw"]
            y_pred = pls_p["y_pred_train"]

            for i in range(X_tr.shape[0]):
                fig_p.add_trace(go.Scatter(
                    x=wn_p, y=X_tr[i], mode="lines",
                    line=dict(width=0.8), opacity=0.5, showlegend=False,
                ), row=1, col=1)

            fig_p.add_trace(go.Scatter(
                x=y_tr, y=y_pred, mode="markers",
                marker=dict(color=COLORS[0], size=7, line=dict(color="black", width=0.5)),
                showlegend=False,
            ), row=1, col=2)
            fig_p.add_trace(go.Scatter(
                x=[y_tr.min(), y_tr.max()], y=[y_tr.min(), y_tr.max()],
                mode="lines", line=dict(dash="dash", color="black", width=1), showlegend=False,
            ), row=1, col=2)

            n_cv = len(pls_p["rmse_cv_all"])
            comps = list(range(1, n_cv + 1))
            fig_p.add_trace(go.Scatter(
                x=comps, y=pls_p["rmse_cv_all"], mode="lines+markers",
                name="CV RMSE", line=dict(color="red"),
            ), row=2, col=1)
            fig_p.add_trace(go.Scatter(
                x=comps, y=pls_p["rmse_train_all"], mode="lines+markers",
                name="Train RMSE", line=dict(color="steelblue"),
            ), row=2, col=1)
            fig_p.add_vline(
                x=pls_p["n_components"], line_dash="dash", line_color="black",
                annotation_text=f"opt={pls_p['n_components']}", row=2, col=1,
            )
            if n_cv > pls_p["n_components"]:
                fig_p.add_vrect(
                    x0=pls_p["n_components"] + 0.5, x1=n_cv + 0.5,
                    fillcolor="rgba(255,0,0,0.12)", line_width=0,
                    annotation_text="Overfitting", annotation_position="top right",
                    annotation=dict(font_size=11, font_color="red", y=0.5, yanchor="middle"),
                    row=2, col=1,
                )

            valid = pls_p["valid_features"]
            wn_valid = wn_p[valid]
            for i in range(pls_p["n_components"]):
                fig_p.add_trace(go.Scatter(
                    x=wn_valid, y=pls_p["model"].x_loadings_[:, i],
                    mode="lines", name=f"PC{i + 1}",
                ), row=2, col=2)

            fig_p.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=1)
            fig_p.update_xaxes(title_text=f"Actual ({unit})",   row=1, col=2)
            fig_p.update_xaxes(title_text="N components",        row=2, col=1)
            fig_p.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=2, col=2)
            fig_p.update_yaxes(title_text="Norm. intensity",    row=1, col=1)
            fig_p.update_yaxes(title_text=f"Predicted ({unit})", row=1, col=2)
            fig_p.update_yaxes(title_text="RMSE",               row=2, col=1)
            fig_p.update_yaxes(title_text="Loading",            row=2, col=2)
            fig_p.update_layout(height=550, legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig_p, use_container_width=True)
            st.caption(
                f"PLS1 calibration for protein concentration prediction using {pls_p['n_components']} latent variables. "
                f"**a)** Overlay of {X_tr.shape[0]} preprocessed standard spectra ({wn_p[0]:.0f}–{wn_p[-1]:.0f} cm⁻¹). "
                f"**b)** Predicted vs actual concentration ({unit}); perfect prediction falls on the dashed diagonal. "
                f"Train R² = {pls_p['r2_train']:.4f}. "
                f"**c)** Leave-one-out CV RMSE (red) and training RMSE (blue) vs number of latent variables; "
                f"dashed line at the selected optimum (opt = {pls_p['n_components']}), "
                f"CV RMSE = {pls_p['cv_rmse']:.4f} {unit}. "
                f"**d)** X-loadings for each latent variable — prominent features reflect the spectral bands driving concentration prediction."
            )

            with st.expander("PLS regression coefficients"):
                coef = _pls_coef(pls_p["model"], len(wn_valid))
                fig_coef = go.Figure()
                fig_coef.add_trace(go.Scatter(
                    x=wn_valid, y=coef[:, 0], mode="lines",
                    name="Protein", line=dict(color=COLORS[0], width=1.5),
                ))
                fig_coef.add_hline(y=0, line_dash="dot", line_color="grey")
                fig_coef.update_layout(
                    xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Coefficient",
                    height=280,
                )
                st.plotly_chart(fig_coef, use_container_width=True)
                st.caption(
                    f"PLS1 regression coefficient vector for protein concentration ({unit}) "
                    f"over {wn_valid[0]:.0f}–{wn_valid[-1]:.0f} cm⁻¹. "
                    f"Positive values indicate spectral features positively correlated with concentration; "
                    f"the magnitude reflects relative contribution to the prediction."
                )

        # ── Salt ─────────────────────────────────────────────
        if pls_s:
            st.divider()
            st.subheader("Salt PLS regression")
            s1, s2, s3 = st.columns(3)
            s1.metric("Optimal components", pls_s["n_components"])
            s2.metric("CV RMSE",  f"{pls_s['cv_rmse']:.4f}")
            s3.metric("Train R²", f"{pls_s['r2_train']:.4f}")

            fig_s = make_subplots(rows=1, cols=3, subplot_titles=[
                "a)  Preprocessed standards", "b)  Calibration (actual vs predicted)", "c)  CV vs Training RMSE"
            ])
            wn_s   = pls_s["wn"]
            X_s    = pls_s["X_train_proc"]
            y_s    = pls_s["y_raw"]
            yp_s   = pls_s["y_pred_train"]

            for i in range(X_s.shape[0]):
                fig_s.add_trace(go.Scatter(
                    x=wn_s, y=X_s[i], mode="lines",
                    line=dict(width=0.8), opacity=0.5, showlegend=False,
                ), row=1, col=1)

            fig_s.add_trace(go.Scatter(
                x=y_s, y=yp_s, mode="markers",
                marker=dict(color=COLORS[3], size=7, line=dict(color="black", width=0.5)),
                showlegend=False,
            ), row=1, col=2)
            fig_s.add_trace(go.Scatter(
                x=[y_s.min(), y_s.max()], y=[y_s.min(), y_s.max()],
                mode="lines", line=dict(dash="dash", color="black", width=1), showlegend=False,
            ), row=1, col=2)

            n_cs = len(pls_s["rmse_cv_all"])
            fig_s.add_trace(go.Scatter(
                x=list(range(1, n_cs + 1)), y=pls_s["rmse_cv_all"],
                mode="lines+markers", name="CV RMSE", line=dict(color="red"),
            ), row=1, col=3)
            fig_s.add_trace(go.Scatter(
                x=list(range(1, n_cs + 1)), y=pls_s["rmse_train_all"],
                mode="lines+markers", name="Train RMSE", line=dict(color="steelblue"),
            ), row=1, col=3)
            fig_s.add_vline(
                x=pls_s["n_components"], line_dash="dash", line_color="black",
                annotation_text=f"opt={pls_s['n_components']}", row=1, col=3,
            )
            if n_cs > pls_s["n_components"]:
                fig_s.add_vrect(
                    x0=pls_s["n_components"] + 0.5, x1=n_cs + 0.5,
                    fillcolor="rgba(255,0,0,0.12)", line_width=0,
                    annotation_text="Overfitting", annotation_position="top right",
                    annotation=dict(font_size=11, font_color="red", y=0.5, yanchor="middle"),
                    row=1, col=3,
                )
            fig_s.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=1)
            fig_s.update_xaxes(title_text="Actual (mM)",        row=1, col=2)
            fig_s.update_xaxes(title_text="Components",         row=1, col=3)
            fig_s.update_yaxes(title_text="Intensity (a.u.)",   row=1, col=1)
            fig_s.update_yaxes(title_text="Predicted (mM)",     row=1, col=2)
            fig_s.update_yaxes(title_text="RMSE (mM)",          row=1, col=3)
            fig_s.update_layout(height=320, legend=dict(orientation="h", y=-0.25))
            st.plotly_chart(fig_s, use_container_width=True)
            st.caption(
                f"PLS1 calibration for salt (sulfate) concentration, built from {X_s.shape[0]} standard spectra "
                f"in the narrow sulfate Raman band ({wn_s[0]:.0f}–{wn_s[-1]:.0f} cm⁻¹) using "
                f"{pls_s['n_components']} latent variable(s). "
                f"**a)** Preprocessed standard spectra. "
                f"**b)** Predicted vs actual salt concentration (mM); train R² = {pls_s['r2_train']:.4f}. "
                f"**c)** Leave-one-out CV RMSE (red) and training RMSE (blue) as a function of latent variables; "
                f"CV RMSE = {pls_s['cv_rmse']:.4f} mM at the selected optimum (opt = {pls_s['n_components']})."
            )

            wn_s_valid = wn_s[pls_s["valid_features"]]
            with st.expander("Salt PLS loadings & coefficients"):
                fig_sl = make_subplots(rows=1, cols=2,
                                       subplot_titles=["a)  Loadings (all PCs)", "b)  Regression coefficients"])
                for i in range(pls_s["n_components"]):
                    fig_sl.add_trace(go.Scatter(
                        x=wn_s_valid, y=pls_s["model"].x_loadings_[:, i],
                        mode="lines", name=f"PC{i + 1}",
                    ), row=1, col=1)
                s_coef = _pls_coef(pls_s["model"], len(wn_s_valid))
                fig_sl.add_trace(go.Scatter(
                    x=wn_s_valid, y=s_coef[:, 0], mode="lines",
                    name="Salt", line=dict(color=COLORS[3], width=1.5), showlegend=False,
                ), row=1, col=2)
                fig_sl.add_hline(y=0, line_dash="dot", line_color="grey", row=1, col=1)
                fig_sl.add_hline(y=0, line_dash="dot", line_color="grey", row=1, col=2)
                fig_sl.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=1)
                fig_sl.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=2)
                fig_sl.update_yaxes(title_text="Loading", row=1, col=1)
                fig_sl.update_yaxes(title_text="Coefficient", row=1, col=2)
                fig_sl.update_layout(height=300, legend=dict(orientation="h", y=-0.3))
                st.plotly_chart(fig_sl, use_container_width=True)
                st.caption(
                    f"**a)** X-loadings for {pls_s['n_components']} latent variable(s) of the salt PLS model, "
                    f"showing the spectral features captured by each PC over "
                    f"{wn_s_valid[0]:.0f}–{wn_s_valid[-1]:.0f} cm⁻¹ (sulfate ν₁ band region). "
                    f"**b)** Regression coefficient vector for salt (mM); "
                    f"the dominant positive feature corresponds to the sulfate stretching mode near 981 cm⁻¹."
                )


# ── Results ───────────────────────────────────────────────────────────────────
with tab_results:
    # ── MCR reference spectra preview (always shown if CSV is uploaded) ──────
    if mcr_init_mode == "PCA from linescans":
        st.info(
            f"MCR initialisation: PCA from linescans. "
            f"ST will be initialised from the top {n_components} principal component(s) of each "
            f"preprocessed linescan at run time. No preview is available until analysis is run."
        )
    elif mcr_ref_src:
        with st.expander("MCR reference spectra (preprocessed)", expanded=False):
            df = _read_csv_src(mcr_ref_src)
            src_label = os.path.basename(mcr_ref_src) if isinstance(mcr_ref_src, str) else mcr_ref_src.name
            comp_names = df.iloc[:, 0].tolist()
            try:
                wn_mcr_raw = pd.to_numeric(df.columns[1:], errors="coerce").values
                X_mcr_raw  = df.iloc[:, 1:].values.astype(float)
                with st.spinner("Preprocessing…"):
                    X_mcr, wn_mcr, _ = an.preprocess_matrix(X_mcr_raw, wn_mcr_raw, settings)
                fig_mcr = go.Figure()
                for i, lbl in enumerate(comp_names):
                    fig_mcr.add_trace(go.Scatter(
                        x=wn_mcr, y=X_mcr[i],
                        mode="lines", name=lbl,
                        line=dict(color=COLORS[i % len(COLORS)], width=1.5),
                    ))
                fig_mcr.update_layout(
                    title="MCR reference spectra", height=300,
                    xaxis_title="Wavenumber (cm⁻¹)",
                    yaxis_title="Intensity (a.u.)",
                    legend=dict(orientation="h", y=-0.3),
                    margin=dict(t=40, b=60),
                )
                st.plotly_chart(fig_mcr, use_container_width=True)
                st.caption(
                    f"MCR reference spectra ({src_label}) used to initialise the MCR-ALS decomposition: "
                    f"{df.shape[0]} component(s) ({comp_names}), "
                    f"preprocessed with the current settings over {wn_mcr[0]:.0f}–{wn_mcr[-1]:.0f} cm⁻¹. "
                    f"These spectra set the initial guess for ST in the alternating least-squares optimisation."
                )
            except Exception as _e:
                st.warning(f"Could not plot MCR references: {_e}")

    if "results" not in st.session_state or not st.session_state["results"]:
        st.info("Run MCR analysis first — configure parameters in the MCR decomposition tab and click ▶ Run MCR Analysis.")
    else:
        results_all = st.session_state["results"]
        models      = st.session_state.get("models", {})
        unit        = st.session_state.get("unit", "mg/mL")
        sm          = st.session_state.get("scan_mode", "xy")
        dist_label  = "Depth (µm)" if sm == "z" else "Distance (µm)"

        pls_p_info  = models.get("pls_protein")
        pls_s_info  = models.get("pls_salt")
        dual        = pls_p_info.get("dual", False) if pls_p_info else False
        comp_labels = models.get("comp_labels", [])

        selected = st.selectbox("Select linescan", list(results_all.keys()))
        r = results_all[selected]

        # ── Row 1: Spectra · MCR profiles · MCR spectra ──────
        fig_top = make_subplots(
            rows=1, cols=3,
            subplot_titles=["a)  Preprocessed spectra", "b)  MCR concentration profiles", "c)  MCR recovered spectra"],
        )
        n_show = min(r["X_proc"].shape[0], 60)
        for i in range(n_show):
            fig_top.add_trace(go.Scatter(
                x=r["wn_proc"], y=r["X_proc"][i], mode="lines",
                line=dict(color="#aaaaaa", width=0.7), opacity=0.4, showlegend=False,
            ), row=1, col=1)

        has_mcr = r["C_mcr"] is not None
        if has_mcr:
            n_comp = r["C_mcr"].shape[1]
            for k in range(n_comp):
                label = comp_labels[k] if k < len(comp_labels) else (
                    COMP_LABELS[k] if k < len(COMP_LABELS) else f"Comp {k + 1}"
                )
                fig_top.add_trace(go.Scatter(
                    x=r["distance"], y=r["C_mcr"][:, k], name=label,
                    line=dict(color=COLORS[k % len(COLORS)], width=1.5),
                ), row=1, col=2)
                fig_top.add_trace(go.Scatter(
                    x=r["wn_proc"], y=r["ST_mcr"][k], name=label,
                    line=dict(color=COLORS[k % len(COLORS)], width=1.5),
                    showlegend=False,
                ), row=1, col=3)
        else:
            fig_top.add_annotation(
                text="No MCR refs uploaded", xref="paper", yref="paper",
                x=0.83, y=0.5, showarrow=False,
            )

        fig_top.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=1)
        fig_top.update_xaxes(title_text=dist_label,           row=1, col=2)
        fig_top.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=3)
        fig_top.update_yaxes(title_text="Intensity (a.u.)",   row=1, col=1)
        fig_top.update_yaxes(title_text="MCR score (a.u.)",   row=1, col=2)
        fig_top.update_yaxes(title_text="Intensity (a.u.)",   row=1, col=3)
        # Legend placed below the concentration profiles subplot (col 2, centre of figure)
        fig_top.update_layout(
            height=340,
            legend=dict(orientation="h", x=0.5, y=-0.22, xanchor="center"),
        )
        st.plotly_chart(fig_top, use_container_width=True)
        if has_mcr:
            _n_mcr_comp = r['C_mcr'].shape[1]
            _mcr_comp_str = ', '.join(
                comp_labels[:_n_mcr_comp] or [f'Comp {k+1}' for k in range(_n_mcr_comp)]
            )
            _cap_b = (
                f"**b)** Spatial MCR-ALS concentration profiles for {_n_mcr_comp} component(s) "
                f"({_mcr_comp_str}) along the linescan — each curve shows how the contribution of that "
                f"component varies with position; MCR-ALS converged in {r['mcr_n_iter']} iteration(s). "
            )
            _cap_c = (
                f"**c)** Corresponding MCR-recovered pure-component spectra (ST matrix rows) — "
                f"these represent the spectral signature attributed to each component by the model."
            )
        else:
            _cap_b = "**b)** MCR concentration profiles — no MCR reference spectra were uploaded, so MCR-ALS was not performed. "
            _cap_c = "**c)** MCR recovered spectra — not available."
        st.caption(
            f"**a)** {n_show} of {r['X_proc'].shape[0]} preprocessed linescan spectra for *{selected}* "
            f"({r['wn_proc'][0]:.0f}–{r['wn_proc'][-1]:.0f} cm⁻¹); individual spectra are overlaid at "
            f"reduced opacity to show the overall spectral envelope across the linescan. "
            + _cap_b + _cap_c
        )

        # ── PCA variance explained (when PCA init was used) ──────────────
        if r.get("pca_var_exp") is not None:
            st.divider()
            with st.expander("PCA initialisation — variance explained"):
                _pve      = r["pca_var_exp"]
                _pld      = r["pca_loadings"]
                _pwn      = r["pca_wn"]
                _pc_lbls  = [f"PC{_k+1}" for _k in range(len(_pve))]
                _cum_pve  = np.cumsum(_pve)
                _fig_pve  = make_subplots(
                    rows=1, cols=2,
                    subplot_titles=["a)  Explained variance (%)", "b)  PCA loadings"],
                )
                _fig_pve.add_trace(go.Bar(
                    x=_pc_lbls, y=_pve, marker_color=COLORS[0],
                    name="Individual", showlegend=True,
                ), row=1, col=1)
                _fig_pve.add_trace(go.Scatter(
                    x=_pc_lbls, y=_cum_pve, mode="lines+markers",
                    name="Cumulative", line=dict(color=COLORS[1], width=2),
                    marker=dict(size=6),
                ), row=1, col=1)
                for _k, _loading in enumerate(_pld):
                    _fig_pve.add_trace(go.Scatter(
                        x=_pwn, y=_loading, mode="lines", name=f"PC{_k+1}",
                        line=dict(color=COLORS[_k % len(COLORS)], width=1.5),
                    ), row=1, col=2)
                _fig_pve.update_yaxes(title_text="Variance (%)", range=[0, 105], row=1, col=1)
                _fig_pve.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=2)
                _fig_pve.update_yaxes(title_text="Loading",             row=1, col=2)
                _fig_pve.update_layout(height=320, legend=dict(orientation="h", y=-0.2))
                st.plotly_chart(_fig_pve, use_container_width=True)
                st.caption(
                    f"PCA computed from the preprocessed linescan spectra of *{selected}* "
                    f"({len(_pve)} component(s)). "
                    f"**a)** Individual (bars) and cumulative (line) explained variance — "
                    f"{len(_pve)} PCs capture {sum(_pve):.1f}% of total variance. "
                    f"**b)** PC loadings used to initialise MCR-ALS."
                )

        # ── MCR summary statistics ────────────────────────────
        if has_mcr:
            st.divider()
            with st.expander("MCR summary"):
                st.markdown(
                    f"**MCR** — converged in {r['mcr_n_iter']} iteration(s), "
                    f"{r['C_mcr'].shape[1]} component(s)."
                )


# ── PLS linescan results (bottom of PLS tab) ──────────────────────────────────
with tab_calib:
    st.divider()
    st.markdown("### Results")
    if "results" not in st.session_state or not st.session_state["results"]:
        st.info("Run analysis first — configure preprocessing above and click ▶ Run MCR Analysis.")
    else:
        _r_all    = st.session_state["results"]
        _models   = st.session_state.get("models", {})
        _unit     = st.session_state.get("unit", "mg/mL")
        _sm       = st.session_state.get("scan_mode", "xy")
        _dl       = "Depth (µm)" if _sm == "z" else "Distance (µm)"
        _pls_p    = _models.get("pls_protein")
        _pls_s    = _models.get("pls_salt")
        _dual     = _pls_p.get("dual", False) if _pls_p else False

        _sel = st.selectbox("Select linescan", list(_r_all.keys()), key="pls_linescan_sel")
        _r   = _r_all[_sel]

        _has_pls  = _pls_p is not None and _r.get("pls_protein") is not None
        _has_peg  = _dual and _r.get("pls_peg") is not None
        _has_salt = _pls_s is not None and _r.get("pls_salt") is not None

        if not _has_pls:
            st.info("No protein PLS model was built — upload a protein standard CSV and click ▶ Build PLS Model.")

        if _has_pls:
            if _dual and _has_peg and _has_salt:
                _specs     = [[{"secondary_y": True}, {"secondary_y": False}]]
                _subtitles = [f"a)  PLS protein ({_unit}) & molecular crowder", "b)  PLS salt (mM)"]
                _n_bot     = 2
            elif _dual and _has_peg:
                _specs     = [[{"secondary_y": True}]]
                _subtitles = [f"a)  PLS protein ({_unit}) & molecular crowder"]
                _n_bot     = 1
            elif _has_salt:
                _specs     = [[{"secondary_y": True}]]
                _subtitles = [f"a)  PLS protein ({_unit}) & salt (mM)"]
                _n_bot     = 1
            else:
                _specs     = [[{"secondary_y": False}]]
                _subtitles = [f"a)  PLS protein ({_unit})"]
                _n_bot     = 1

            _c1sy    = _specs[0][0].get("secondary_y", False)
            _fig_bot = make_subplots(rows=1, cols=_n_bot, specs=_specs, subplot_titles=_subtitles)
            _dist    = _r["distance"]
            _cv_p    = _pls_p["cv_rmse_protein"] if _dual else _pls_p["cv_rmse"]
            _prot    = _r["pls_protein"]

            _fig_bot.add_trace(go.Scatter(
                x=_dist, y=_prot, mode="lines",
                line=dict(color=COLORS[0], width=1.5), showlegend=False,
            ), row=1, col=1, **({} if not _c1sy else {"secondary_y": False}))
            _fig_bot.add_trace(go.Scatter(
                x=np.concatenate([_dist, _dist[::-1]]),
                y=np.concatenate([_prot + _cv_p, (_prot - _cv_p)[::-1]]),
                fill="toself", fillcolor=COLORS_FILL[0],
                line=dict(color="rgba(0,0,0,0)"), showlegend=False, hoverinfo="skip",
            ), row=1, col=1, **({} if not _c1sy else {"secondary_y": False}))
            _fig_bot.update_yaxes(
                title_text=f"Protein ({_unit})",
                title_font=dict(color=COLORS[0]), tickfont=dict(color=COLORS[0]),
                row=1, col=1, **({} if not _c1sy else {"secondary_y": False}),
            )

            if _dual and _has_peg:
                _cv_peg = _pls_p["cv_rmse_peg"]
                _peg    = _r["pls_peg"]
                _fig_bot.add_trace(go.Scatter(
                    x=_dist, y=_peg, mode="lines",
                    line=dict(color=COLORS[2], width=1.5), showlegend=False,
                ), row=1, col=1, secondary_y=True)
                _fig_bot.add_trace(go.Scatter(
                    x=np.concatenate([_dist, _dist[::-1]]),
                    y=np.concatenate([_peg + _cv_peg, (_peg - _cv_peg)[::-1]]),
                    fill="toself", fillcolor=COLORS_FILL[2],
                    line=dict(color="rgba(0,0,0,0)"), showlegend=False, hoverinfo="skip",
                ), row=1, col=1, secondary_y=True)
                _fig_bot.update_yaxes(
                    title_text="Molecular crowder (wt%)",
                    title_font=dict(color=COLORS[2]), tickfont=dict(color=COLORS[2]),
                    row=1, col=1, secondary_y=True,
                )

            if _has_salt:
                _s_col = 2 if (_dual and _has_peg) else 1
                _s_sy  = True if _s_col == 1 else None
                _cv_s  = _pls_s["cv_rmse"]
                _salt  = _r["pls_salt"]
                _s_kw  = {"secondary_y": _s_sy} if _s_sy is not None else {}
                _fig_bot.add_trace(go.Scatter(
                    x=_dist, y=_salt, mode="lines",
                    line=dict(color=COLORS[3], width=1.5), showlegend=False,
                ), row=1, col=_s_col, **_s_kw)
                _fig_bot.add_trace(go.Scatter(
                    x=np.concatenate([_dist, _dist[::-1]]),
                    y=np.concatenate([_salt + _cv_s, (_salt - _cv_s)[::-1]]),
                    fill="toself", fillcolor=COLORS_FILL[3],
                    line=dict(color="rgba(0,0,0,0)"), showlegend=False, hoverinfo="skip",
                ), row=1, col=_s_col, **_s_kw)
                _fig_bot.update_yaxes(
                    title_text="Salt (mM)",
                    title_font=dict(color=COLORS[3]), tickfont=dict(color=COLORS[3]),
                    row=1, col=_s_col, **_s_kw,
                )

            for _col in range(1, _n_bot + 1):
                _fig_bot.update_xaxes(title_text=_dl, row=1, col=_col)
            _fig_bot.update_layout(height=320)
            st.plotly_chart(_fig_bot, use_container_width=True)
            _pls_parts = [f"protein ± {_cv_p:.4f} {_unit}"]
            if _has_peg:
                _pls_parts.append(f"molecular crowder ± {_pls_p['cv_rmse_peg']:.4f} wt%")
            if _has_salt:
                _pls_parts.append(f"salt ± {_pls_s['cv_rmse']:.4f} mM")
            st.caption(
                f"**a)** PLS-predicted concentration profiles along the linescan for *{_sel}*. "
                f"Shaded bands indicate ±CV RMSE: {'; '.join(_pls_parts)}. "
                f"Left y-axis: protein; right y-axis: second analyte where applicable."
                + (f" **b)** Salt (mM) on a separate axis." if (_dual and _has_peg and _has_salt) else "")
            )

            if _dual and _has_peg:
                st.divider()
                _fig_sc = go.Figure()
                _fig_sc.add_trace(go.Scatter(
                    x=_r["pls_protein"], y=_r["pls_peg"], mode="markers",
                    marker=dict(color=COLORS[2], size=7, line=dict(color="black", width=0.5)),
                    error_x=dict(type="constant", value=_pls_p["cv_rmse_protein"],
                                 color="black", thickness=1, width=4),
                    error_y=dict(type="constant", value=_pls_p["cv_rmse_peg"],
                                 color="black", thickness=1, width=4),
                ))
                _fig_sc.update_layout(
                    title="Molecular crowder vs Protein (PLS)",
                    xaxis_title=f"Protein ({_unit})", yaxis_title="Molecular crowder (wt%)",
                    height=350, width=420,
                )
                st.plotly_chart(_fig_sc)
                st.caption(
                    f"Molecular crowder vs protein co-localisation scatter for *{_sel}*. "
                    f"Error bars: ±CV RMSE (protein: {_pls_p['cv_rmse_protein']:.4f} {_unit}; "
                    f"molecular crowder: {_pls_p['cv_rmse_peg']:.4f} wt%)."
                )
            elif not _dual and _has_salt:
                st.divider()
                _fig_sc = go.Figure()
                _fig_sc.add_trace(go.Scatter(
                    x=_r["pls_protein"], y=_r["pls_salt"], mode="markers",
                    marker=dict(color=COLORS[2], size=7, line=dict(color="black", width=0.5)),
                    error_x=dict(type="constant", value=_pls_p["cv_rmse"],
                                 color="black", thickness=1, width=4),
                    error_y=dict(type="constant", value=_pls_s["cv_rmse"],
                                 color="black", thickness=1, width=4),
                ))
                _fig_sc.update_layout(
                    title="Salt vs Protein (PLS)",
                    xaxis_title=f"Protein ({_unit})", yaxis_title="Salt (mM)",
                    height=350, width=420,
                )
                st.plotly_chart(_fig_sc)
                st.caption(
                    f"Salt vs protein co-localisation scatter for *{_sel}*. "
                    f"Error bars: ±CV RMSE (protein: {_pls_p['cv_rmse']:.4f} {_unit}; "
                    f"salt: {_pls_s['cv_rmse']:.4f} mM)."
                )

        st.divider()
        with st.expander("Summary statistics"):
            _n_sc = (1 if _has_pls else 0) + (1 if _has_peg else 0) + (1 if _has_salt else 0)
            _cols = st.columns(max(_n_sc, 1))
            _ci   = 0
            if _has_pls:
                _cv_p_val = _pls_p["cv_rmse_protein"] if _dual else _pls_p["cv_rmse"]
                _cols[_ci].markdown(
                    f"**Protein (PLS)**  \n"
                    f"Min: {_r['pls_protein'].min():.3f} {_unit}  \n"
                    f"Max: {_r['pls_protein'].max():.3f} {_unit}  \n"
                    f"Mean: {_r['pls_protein'].mean():.3f} {_unit}  \n"
                    f"±CV RMSE: {_cv_p_val:.3f} {_unit}"
                )
                _ci += 1
            if _has_peg:
                _cols[_ci].markdown(
                    f"**Molecular crowder (PLS)**  \n"
                    f"Min: {_r['pls_peg'].min():.3f} wt%  \n"
                    f"Max: {_r['pls_peg'].max():.3f} wt%  \n"
                    f"Mean: {_r['pls_peg'].mean():.3f} wt%  \n"
                    f"±CV RMSE: {_pls_p['cv_rmse_peg']:.3f} wt%"
                )
                _ci += 1
            if _has_salt:
                _cols[_ci].markdown(
                    f"**Salt (PLS)**  \n"
                    f"Min: {_r['pls_salt'].min():.3f} mM  \n"
                    f"Max: {_r['pls_salt'].max():.3f} mM  \n"
                    f"Mean: {_r['pls_salt'].mean():.3f} mM  \n"
                    f"±CV RMSE: {_pls_s['cv_rmse']:.3f} mM"
                )


# ── Further analysis ──────────────────────────────────────────────────────────
with tab_further:
    if not linescan_files:
        st.info("Upload linescan files in the Files tab to enable further analysis.")
    else:
        from sklearn.decomposition import PCA as _SklearnPCA
        from scipy.optimize import curve_fit as _curve_fit

        # ── Drag-and-drop pipeline helpers ────────────────────────────────────
        # Presets: name → (steps_list, seed_state_dict or None)
        # seed_state keys are the item label (e.g. "Spectral cut #2") and map param → value.
        _PIPE_PRESETS = {
            "— custom —": None,
            "Amide I band decomposition": (
                ["spectral_cut","spike_removal","spectral_cut","savgol","area","endpoint"],
                {"Spectral cut #2": {"wn_min": 1580, "wn_max": 1720, "use_gap": False}},
            ),
            "General (rubberband + SNV)":   (["spectral_cut","spike_removal","rubberband","snv"], None),
            "PCA (rubberband + SNV)":        (["spectral_cut","spike_removal","rubberband","snv"], None),
            "Peak ratio (rubberband only)":  (["spectral_cut","spike_removal","rubberband"], None),
        }
        _PIPE_CATS = [
            ("Spectral cut",         ["spectral_cut"]),
            ("Spike removal",        ["spike_removal"]),
            ("Baseline subtraction", ["rubberband","rolling_ball","als","arpls","airpls",
                                       "iasls","drpls","imodpoly","modpoly","poly","endpoint","linear"]),
            ("Smoothing",            ["savgol","gaussian","fft_lowpass"]),
            ("Normalisation",        ["snv","area","vector","minmax"]),
        ]
        _PKLAB = {
            "spectral_cut":"Spectral cut", "spike_removal":"Spike removal",
            "rubberband":"Rubberband",     "rolling_ball":"Rolling ball",
            "als":"ALS",      "arpls":"ARPLS",   "airpls":"AIRPLS",
            "iasls":"IASLS",  "drpls":"DRPLS",   "imodpoly":"IModPoly",
            "modpoly":"ModPoly","poly":"Poly",    "endpoint":"Endpoint",
            "linear":"Linear", "savgol":"Savitzky-Golay",
            "gaussian":"Gaussian","fft_lowpass":"FFT low-pass",
            "snv":"SNV","area":"Area","vector":"Vector","minmax":"Min-max",
        }
        _PKLKEY  = {v: k for k, v in _PKLAB.items()}
        _NO_PPAR = {"spike_removal","rubberband","endpoint","linear",
                    "snv","area","vector","minmax"}

        def _pkey(s):
            """'Rubberband #2' → 'rubberband'"""
            return _PKLKEY.get(s.rsplit(" #", 1)[0], "")

        def _wk(pfx, s, p):
            _safe = s.replace(" ","_").replace("#","n").replace(".","_").replace("-","_")
            return f"pp_{pfx}_{_safe}_{p}"

        def _pipe_init(pfx, defaults, seed_state=None):
            """Build pipeline from defaults list.
            seed_state: optional dict mapping item name → {param: value} to pre-fill widgets.
            """
            _ctr, _items = {}, []
            for _k in defaults:
                _lbl = _PKLAB.get(_k, _k)
                _n   = _ctr.get(_lbl, 1)
                _items.append(f"{_lbl} #{_n}")
                _ctr[_lbl] = _n + 1
            st.session_state[f"{pfx}_pipe"]     = _items
            st.session_state[f"{pfx}_ctr"]      = _ctr
            st.session_state[f"{pfx}_sort_ver"] = 0
            if seed_state:
                for _item, _params in seed_state.items():
                    for _par, _val in _params.items():
                        st.session_state[_wk(pfx, _item, _par)] = _val

        def _step_params_ui(pfx, item):
            """Render parameter widgets for one pipeline item."""
            _sk = _pkey(item)
            if _sk == "spectral_cut":
                _c1, _c2 = st.columns(2)
                _c1.number_input("Min (cm⁻¹)", value=700,  step=50, key=_wk(pfx,item,"wn_min"))
                _c2.number_input("Max (cm⁻¹)", value=3900, step=50, key=_wk(pfx,item,"wn_max"))
                st.toggle("Exclude gap region", value=True, key=_wk(pfx,item,"use_gap"))
                if st.session_state.get(_wk(pfx,item,"use_gap"), True):
                    _g1, _g2 = st.columns(2)
                    _g1.number_input("Gap from", value=1850, step=50, key=_wk(pfx,item,"gap_lo"))
                    _g2.number_input("Gap to",   value=2750, step=50, key=_wk(pfx,item,"gap_hi"))
            elif _sk == "rolling_ball":
                st.slider("Ball radius", 10, 300, 50, key=_wk(pfx,item,"radius"))
            elif _sk == "als":
                _c1, _c2 = st.columns(2)
                _c1.number_input("λ (smoothness)", value=1e5, min_value=1e2, max_value=1e8,
                                  format="%.0e", key=_wk(pfx,item,"lam"))
                _c2.number_input("p (asymmetry)",  value=0.01, min_value=0.001, max_value=0.5,
                                  format="%.3f",  key=_wk(pfx,item,"p"))
            elif _sk in ("arpls","airpls","iasls","drpls"):
                st.number_input("λ (smoothness)", value=1e5, min_value=1e2, max_value=1e9,
                                format="%.0e", key=_wk(pfx,item,"lam"))
            elif _sk in ("imodpoly","modpoly","poly"):
                st.slider("Polynomial order", 1, 8, 2, key=_wk(pfx,item,"order"))
            elif _sk == "savgol":
                _c1, _c2 = st.columns(2)
                _c1.slider("Window length (odd)", 5, 51, 11, step=2, key=_wk(pfx,item,"window"))
                _c2.slider("Poly order", 1, 9, 3, key=_wk(pfx,item,"poly"))
            elif _sk == "gaussian":
                st.slider("σ (sigma)", 0.5, 5.0, 1.0, step=0.5, key=_wk(pfx,item,"sigma"))
            elif _sk == "fft_lowpass":
                st.slider("Cutoff fraction", 0.01, 0.5, 0.1, step=0.01, key=_wk(pfx,item,"cutoff"))

        def _pipeline_ui(pfx, defaults, seed_defaults=None):
            """Render pipeline builder. Returns ordered list of item strings."""
            _ss  = f"{pfx}_pipe"
            _ssc = f"{pfx}_ctr"
            if _ss not in st.session_state:
                _pipe_init(pfx, defaults, seed_state=seed_defaults)

            # ── Preset selector ───────────────────────────────────────────
            _ph1, _ph2, _ph3 = st.columns([3, 1, 1])
            _preset_sel = _ph1.selectbox(
                "Load preset", list(_PIPE_PRESETS.keys()),
                key=f"{pfx}_preset_sel",
                label_visibility="collapsed",
            )
            if _ph2.button("Load", key=f"{pfx}_preset_load",
                           help="Replace current pipeline with selected preset"):
                _pval = _PIPE_PRESETS[_preset_sel]
                if _pval is not None:
                    _psteps, _pseed = _pval
                    _pipe_init(pfx, _psteps, seed_state=_pseed)
                    st.rerun()
            if _ph3.button("↺ Reset", key=f"{pfx}_rst", help="Restore default pipeline"):
                _pipe_init(pfx, defaults, seed_state=seed_defaults)
                st.rerun()

            _ph1.caption(
                "Click **+** to add steps. Drag to reorder. **×** to remove."
            )

            _lc, _rc = st.columns([1, 2])

            # ── Palette (left) ────────────────────────────────────────────
            with _lc:
                st.markdown("**Available steps**")
                for _cat, _ckeys in _PIPE_CATS:
                    with st.expander(_cat, expanded=(_cat != "Baseline subtraction")):
                        for _k in _ckeys:
                            _lbl = _PKLAB[_k]
                            if st.button(f"+ {_lbl}", key=f"{pfx}_add_{_k}",
                                         use_container_width=True):
                                _c = st.session_state[_ssc]
                                _n = _c.get(_lbl, 1)
                                st.session_state[_ss].append(f"{_lbl} #{_n}")
                                _c[_lbl] = _n + 1
                                st.session_state[f"{pfx}_sort_ver"] = (
                                    st.session_state.get(f"{pfx}_sort_ver", 0) + 1
                                )
                                st.rerun()

            # ── Pipeline (right) ──────────────────────────────────────────
            with _rc:
                _cur = list(st.session_state[_ss])
                if not _cur:
                    st.caption("_No steps added yet — click + on the left._")
                else:
                    st.markdown("**Pipeline** (drag to reorder):")
                    try:
                        from streamlit_sortables import sort_items as _si
                        _sort_ver = st.session_state.get(f"{pfx}_sort_ver", 0)
                        _new = _si(_cur, key=f"{pfx}_sortable_{_sort_ver}")
                        if list(_new) != _cur:
                            st.session_state[_ss] = list(_new)
                    except ImportError:
                        st.info("Install `streamlit-sortables` for drag-and-drop reordering.")

                    st.markdown("**Step configuration:**")
                    for _i, _item in enumerate(list(st.session_state[_ss])):
                        _sk = _pkey(_item)
                        _pc, _rx = st.columns([11, 1])
                        with _rx:
                            if st.button("×", key=f"{pfx}_rm_{_i}",
                                          help=f"Remove {_item}"):
                                st.session_state[_ss].pop(_i)
                                st.session_state[f"{pfx}_sort_ver"] = (
                                    st.session_state.get(f"{pfx}_sort_ver", 0) + 1
                                )
                                st.rerun()
                        with _pc:
                            if _sk not in _NO_PPAR and _sk:
                                with st.expander(f"⚙ {_item}", expanded=False):
                                    _step_params_ui(pfx, _item)
                            else:
                                st.caption(f"✓ {_item}  _(no parameters)_")

            return list(st.session_state[_ss])

        def _apply_pipeline(X, wn, items, pfx, fine_lo=None, fine_hi=None):
            """Execute pipeline steps in order; apply optional fine wavenumber cut at the end.

            When fine_lo/fine_hi are given, any trailing endpoint/linear steps in the
            pipeline are deferred until after the fine cut so they anchor the endpoints
            of the final spectral window (not the wider pre-cut window) to zero.
            """
            X  = X.copy().astype(float)
            wn = wn.copy().astype(float)

            def _p(item, param, default):
                return st.session_state.get(_wk(pfx, item, param), default)

            # Collect trailing endpoint/linear steps to run after the fine cut
            _main  = list(items)
            _deferred = []
            if fine_lo is not None or fine_hi is not None:
                while _main and _pkey(_main[-1]) in ("endpoint", "linear"):
                    _deferred.insert(0, _main.pop())

            for _item in _main:
                _sk = _pkey(_item)
                if not _sk or X.shape[1] == 0:
                    continue

                if _sk == "spectral_cut":
                    _lo, _hi = _p(_item,"wn_min",700.0), _p(_item,"wn_max",3900.0)
                    _mask = (wn >= _lo) & (wn <= _hi)
                    if _p(_item,"use_gap",True):
                        _glo, _ghi = _p(_item,"gap_lo",1850.0), _p(_item,"gap_hi",2750.0)
                        if _glo is not None and _ghi is not None:
                            _mask &= ~((wn >= _glo) & (wn <= _ghi))
                    X, wn = X[:,_mask], wn[_mask]

                elif _sk == "spike_removal":
                    for _i in range(X.shape[0]):
                        X[_i] = rms.spike_removal_scp(X[_i])

                elif _sk == "rubberband":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.rubberband_correction(wn, X[_i])

                elif _sk == "rolling_ball":
                    _r = int(_p(_item,"radius",50))
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.rolling_ball_baseline(X[_i], ball_radius=_r)

                elif _sk == "als":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.als_baseline_correction(
                            X[_i], lam=_p(_item,"lam",1e5), p=_p(_item,"p",0.01))

                elif _sk in ("arpls","airpls","iasls","drpls"):
                    import ramanspy.preprocessing.baseline as _rsb
                    _fn = {"arpls":_rsb._arpls,"airpls":_rsb._airpls,
                           "iasls":_rsb._iasls,"drpls":_rsb._drpls}[_sk]
                    _lam = _p(_item,"lam",1e5)
                    for _i in range(X.shape[0]):
                        _res, _ = _fn(X[_i:_i+1], wn, lam=_lam)
                        X[_i] -= _res.squeeze()

                elif _sk in ("imodpoly","modpoly","poly"):
                    import ramanspy.preprocessing.baseline as _rsb
                    _fn = {"imodpoly":_rsb._imodpoly,"modpoly":_rsb._modpoly,
                           "poly":_rsb._poly}[_sk]
                    _ord = int(_p(_item,"order",2))
                    for _i in range(X.shape[0]):
                        _res, _ = _fn(X[_i:_i+1], wn, poly_order=_ord)
                        X[_i] -= _res.squeeze()

                elif _sk == "endpoint":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.endpoint_baseline(X[_i])

                elif _sk == "linear":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.linear_baseline(X[_i])

                elif _sk == "savgol":
                    _win  = int(_p(_item,"window",11))
                    _poly = int(_p(_item,"poly",3))
                    if _win > _poly and _win % 2 == 1:
                        for _i in range(X.shape[0]):
                            X[_i] = rms.Savgol_filter(X[_i], _win, _poly)

                elif _sk == "gaussian":
                    from scipy.ndimage import gaussian_filter as _gfilt
                    _sig = float(_p(_item,"sigma",1.0))
                    for _i in range(X.shape[0]):
                        X[_i] = _gfilt(X[_i], sigma=_sig)

                elif _sk == "fft_lowpass":
                    _cut = float(_p(_item,"cutoff",0.1))
                    _n   = X.shape[1]
                    for _i in range(X.shape[0]):
                        _F = np.fft.rfft(X[_i])
                        _F[max(1,int(_cut*len(_F))):] = 0
                        X[_i] = np.fft.irfft(_F, n=_n)

                elif _sk == "snv":
                    X = rms.snv_normalization(X)
                elif _sk == "area":
                    X = rms.area_normalization(X, wn)
                elif _sk == "vector":
                    X = rms.vector_normalization(X)
                elif _sk == "minmax":
                    _mn, _mx = X.min(axis=1,keepdims=True), X.max(axis=1,keepdims=True)
                    X = np.where(_mx-_mn > 0, (X-_mn)/(_mx-_mn), 0.0)

            # Fine cut to analysis-specific region (e.g. amide I 1600–1700 cm⁻¹).
            # Applied BEFORE any trailing endpoint/linear steps so those steps
            # anchor the endpoints of the final spectral window to zero.
            if X.shape[1] > 0 and (fine_lo is not None or fine_hi is not None):
                _lo  = fine_lo  if fine_lo  is not None else float(wn.min())
                _hi  = fine_hi  if fine_hi  is not None else float(wn.max())
                _m   = (wn >= _lo) & (wn <= _hi)
                X, wn = X[:,_m], wn[_m]

            # Run any trailing endpoint / linear steps that were deferred
            for _item in _deferred:
                _sk = _pkey(_item)
                if _sk == "endpoint":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.endpoint_baseline(X[_i])
                elif _sk == "linear":
                    for _i in range(X.shape[0]):
                        X[_i] -= rms.linear_baseline(X[_i])

            return X, wn

        results_all = st.session_state.get("results", {})
        models      = st.session_state.get("models", {})
        unit        = st.session_state.get("unit", "mg/mL")
        sm          = st.session_state.get("scan_mode", "xy")
        dist_label  = "Depth (µm)" if sm == "z" else "Distance (µm)"
        pls_p_info  = models.get("pls_protein")
        dual        = pls_p_info.get("dual", False) if pls_p_info else False
        comp_labels = models.get("comp_labels", [])

        # ── Dataset selector ───────────────────────────────────────────────
        _all_fnames = [_f.name for _f in linescan_files]
        _sel_fnames = st.multiselect(
            "Datasets to include",
            options=_all_fnames,
            default=_all_fnames,
            key="further_datasets",
            help="Select one or more linescans to include in all analyses below.",
        )
        if not _sel_fnames:
            st.info("Select at least one dataset above.")
            st.stop()

        # ── Load raw files; also retain raw spectra for per-analysis preprocessing ─
        _X_list, _X_raw_list, _dist_list, _fname_list = [], [], [], []
        _prot_list, _peg_list, _salt_list, _mcr_list = [], [], [], []
        _wn = None
        _wn_raw_shared = None  # raw wavenumber axis for fresh preprocessing
        for _f in linescan_files:
            if _f.name not in _sel_fnames:
                continue
            _f.seek(0)
            _wn_raw, _X_raw, _pos = an.load_linescan_bytes(_f.read(), _f.name)
            _f.seek(0)
            _X_proc_f, _wn_proc_f, _ = an.preprocess_matrix(_X_raw, _wn_raw, settings)
            _dist_f = an.compute_cumulative_distance(_pos, an.detect_scan_mode(_pos))
            _n = _X_proc_f.shape[0]
            _X_list.append(_X_proc_f)
            _X_raw_list.append(_X_raw)
            _dist_list.extend(_dist_f.tolist())
            _fname_list.extend([_f.name] * _n)
            if _wn is None:
                _wn = _wn_proc_f
            if _wn_raw_shared is None:
                _wn_raw_shared = _wn_raw
            # PLS/MCR scores from cached results (for colouring), if available
            _r = results_all.get(_f.name, {})
            if _r.get("pls_protein") is not None:
                _prot_list.extend(_r["pls_protein"].tolist())
            else:
                _prot_list.extend([np.nan] * _n)
            if dual and _r.get("pls_peg") is not None:
                _peg_list.extend(_r["pls_peg"].tolist())
            if _r.get("pls_salt") is not None:
                _salt_list.extend(_r["pls_salt"].tolist())
            if _r.get("C_mcr") is not None:
                _mcr_list.append(_r["C_mcr"])

        if not _X_list:
            st.error("No spectra could be loaded. Check that linescan files are valid.")
            st.stop()
        _X_all    = np.vstack(_X_list)
        _X_raw_all = np.vstack(_X_raw_list)  # raw spectra for per-analysis preprocessing
        _prot_raw = np.array(_prot_list)
        _has_pls_results = not np.all(np.isnan(_prot_raw))
        _prot   = _prot_raw if _has_pls_results else None
        _peg    = np.array(_peg_list)   if _peg_list   else None
        _salt   = np.array(_salt_list)  if _salt_list  else None
        _mcr_C  = np.vstack(_mcr_list)  if _mcr_list   else None

        st.divider()

        # ── Global concentration filter (only when PLS results exist) ─────
        _global_conc_lo = _global_conc_hi = _global_conc = None
        if _has_pls_results:
            _prot_min_g = float(np.nanmin(_prot))
            _prot_max_g = float(np.nanmax(_prot))
            _prot_step  = max((_prot_max_g - _prot_min_g) / 200, 1e-6)
            _global_conc_range = st.slider(
                f"Protein concentration range ({unit})",
                min_value=_prot_min_g, max_value=_prot_max_g,
                value=(_prot_min_g, _prot_max_g),
                step=_prot_step,
                key="global_conc_thresh",
                help="Keep only spectra whose PLS protein concentration falls within this range. Applied to all three analyses — PCA, amide decomposition, and peak ratio.",
            )
            _global_conc_lo, _global_conc_hi = _global_conc_range
            _global_conc = _global_conc_lo   # backward-compat alias used in amide average
            _gmask      = (_prot >= _global_conc_lo) & (_prot <= _global_conc_hi)
            _gidx       = np.where(_gmask)[0]
            _X_all      = _X_all[_gmask]
            _X_raw_all  = _X_raw_all[_gmask]
            _prot       = _prot[_gmask]
            _peg        = _peg[_gmask]    if _peg   is not None else None
            _salt       = _salt[_gmask]   if _salt  is not None else None
            _mcr_C      = _mcr_C[_gmask]  if _mcr_C is not None else None
            _dist_list  = [_dist_list[_i]  for _i in _gidx]
            _fname_list = [_fname_list[_i] for _i in _gidx]
        else:
            st.info("No PLS protein model was built — concentration filtering is disabled. Upload a protein standard CSV and click ▶ Build Models to enable it.")

        # ── MCR score range filter ─────────────────────────────────────────
        if _mcr_C is not None and _mcr_C.shape[1] > 0:
            _mcr_comp_labels = [
                comp_labels[_k] if _k < len(comp_labels) else f"MCR comp {_k+1}"
                for _k in range(_mcr_C.shape[1])
            ]
            _mf1, _mf2 = st.columns([1, 3])
            _mcr_filt_comp = _mf1.selectbox(
                "MCR component", _mcr_comp_labels, key="mcr_filt_comp",
                help="Choose which MCR score to filter on.",
            )
            _mcr_filt_idx  = _mcr_comp_labels.index(_mcr_filt_comp)
            _mcr_col       = _mcr_C[:, _mcr_filt_idx]
            _mcr_lo, _mcr_hi = float(_mcr_col.min()), float(_mcr_col.max())
            _mcr_step = max((_mcr_hi - _mcr_lo) / 200, 1e-6)
            _mcr_range = _mf2.slider(
                f"MCR score range — {_mcr_filt_comp}",
                min_value=_mcr_lo, max_value=_mcr_hi,
                value=(_mcr_lo, _mcr_hi),
                step=_mcr_step,
                key="mcr_score_range",
                help="Keep only spectra whose MCR score for the selected component falls within this range.",
            )
            _mmask      = (_mcr_col >= _mcr_range[0]) & (_mcr_col <= _mcr_range[1])
            _midx       = np.where(_mmask)[0]
            _X_all      = _X_all[_mmask]
            _X_raw_all  = _X_raw_all[_mmask]
            _prot       = _prot[_mmask]  if _prot  is not None else None
            _peg        = _peg[_mmask]   if _peg   is not None else None
            _salt       = _salt[_mmask]  if _salt  is not None else None
            _mcr_C      = _mcr_C[_mmask]
            _dist_list  = [_dist_list[_i]  for _i in _midx]
            _fname_list = [_fname_list[_i] for _i in _midx]

        # ── Label options (always include position; add PLS/MCR if available) ─
        _label_opts = {}
        _label_opts["Distance (µm)"]  = np.array(_dist_list, dtype=float)
        _label_opts["Spectrum index"] = np.arange(_X_all.shape[0], dtype=float)
        if _has_pls_results and _prot is not None:
            _label_opts[f"PLS Protein ({unit})"] = _prot
        if _peg  is not None: _label_opts["PLS Molecular crowder (wt%)"] = _peg
        if _salt is not None: _label_opts["PLS Salt"]       = _salt
        if _mcr_C is not None:
            for _k in range(_mcr_C.shape[1]):
                _lbl = comp_labels[_k] if _k < len(comp_labels) else f"MCR comp {_k+1}"
                _label_opts[f"MCR: {_lbl}"] = _mcr_C[:, _k]

        _filter_desc = f"dataset selection"
        if _has_pls_results and _global_conc_lo is not None:
            _filter_desc += f", concentration range [{_global_conc_lo:.2f}–{_global_conc_hi:.2f}] {unit}"
        if _mcr_C is not None:
            _filter_desc += ", MCR score range"
        _fig_caption(
            f"{_X_all.shape[0]} spectra pass all active filters "
            f"across {len(results_all)} linescan(s). "
            f"Filters: {_filter_desc}."
        )

        # ── a. PCA ────────────────────────────────────────────────────────
        st.markdown("### a. PCA of a selected spectral region")
        _pca3, _pca4 = st.columns(2)
        _pca_n_comp  = _pca3.slider("Number of PCs", 2, 10, 3, key="pca_n_comp")
        _pca_lbl_key = _pca4.selectbox("Colour by", list(_label_opts.keys()), key="pca_label")
        _items_pca = _pipeline_ui(
            "pca", ["spectral_cut", "spike_removal", "rubberband", "snv"])
        _run_pca = st.button("▶ Run PCA", key="btn_pca", on_click=_goto_further)

        if _run_pca:
            with st.spinner("Preprocessing spectra for PCA…"):
                _X_proc, _wn2 = _apply_pipeline(
                    _X_raw_all, _wn_raw_shared, _items_pca, "pca",
                )
            if _X_proc.shape[1] == 0:
                st.error(
                    "No wavenumber points remain after preprocessing. "
                    "Check the Spectral cut step parameters in the pipeline."
                )
            else:
                if _X_proc.shape[1] < 2:
                    st.error("Analysis cut contains fewer than 2 wavenumber points after preprocessing.")
                else:
                    _pca_model = _SklearnPCA(
                        n_components=min(_pca_n_comp, _X_proc.shape[1], _X_proc.shape[0])
                    )
                    _scores = _pca_model.fit_transform(_X_proc)
                    st.session_state["pca_result"] = dict(
                        scores=_scores,
                        loadings=_pca_model.components_,
                        var_exp=_pca_model.explained_variance_ratio_ * 100,
                        wn=_wn2,
                        label_key=_pca_lbl_key,
                        label_vals={k: v.tolist() for k, v in _label_opts.items()},
                    )

        if "pca_result" in st.session_state:
            _pr      = st.session_state["pca_result"]
            _n_pc    = _pr["scores"].shape[1]
            _var_exp = _pr["var_exp"]

            # Build axis options: all PCs + all stored labels
            _pc_opts  = [f"PC{_i+1} ({_var_exp[_i]:.1f} %)" for _i in range(_n_pc)]
            _lbl_opts_stored = list(_pr.get("label_vals", {}).keys())
            _ax_opts  = _pc_opts + _lbl_opts_stored

            def _get_axis(key):
                if key.startswith("PC"):
                    _idx = int(key.split(" ")[0][2:]) - 1
                    return _pr["scores"][:, _idx], key
                else:
                    return np.array(_pr["label_vals"][key]), key

            # Default: X = PLS protein, Y = PC1, colour = PLS protein
            _prot_idx = len(_pc_opts)   # protein label sits after all PC options
            _prot_idx = min(_prot_idx, len(_ax_opts) - 1)
            _axcol1, _axcol2, _axcol3 = st.columns(3)
            _x_key = _axcol1.selectbox("X axis",    _ax_opts, index=_prot_idx, key="pca_x_axis2")
            _y_key = _axcol2.selectbox("Y axis",    _ax_opts, index=0,          key="pca_y_axis2")
            _c_key = _axcol3.selectbox("Colour by", _ax_opts, index=_prot_idx, key="pca_color2")

            _x_vals, _x_lbl = _get_axis(_x_key)
            _y_vals, _y_lbl = _get_axis(_y_key)
            _c_vals, _c_lbl = _get_axis(_c_key)

            _fig_pca  = make_subplots(
                rows=1, cols=3,
                subplot_titles=[f"a)  Scores ({_x_key.split(' ')[0]} vs {_y_key.split(' ')[0]})",
                                 "b)  Loadings", "c)  Explained variance (%)"],
            )
            _fig_pca.add_trace(go.Scatter(
                x=_x_vals, y=_y_vals, mode="markers",
                marker=dict(
                    color=_c_vals, colorscale="Viridis", size=5,
                    colorbar=dict(title=_c_lbl, thickness=12, len=0.7),
                    showscale=True,
                ),
                showlegend=False,
            ), row=1, col=1)
            for _i, _loading in enumerate(_pr["loadings"]):
                _fig_pca.add_trace(go.Scatter(
                    x=_pr["wn"], y=_loading, mode="lines", name=f"PC{_i+1}",
                    line=dict(color=COLORS[_i % len(COLORS)], width=1.5),
                ), row=1, col=2)
            _pc_labels = [f"PC{_i+1}" for _i in range(len(_var_exp))]
            _cum_var   = np.cumsum(_var_exp)
            _fig_pca.add_trace(go.Bar(
                x=_pc_labels, y=_var_exp,
                marker_color=COLORS[0], name="Individual", showlegend=True,
            ), row=1, col=3)
            _fig_pca.add_trace(go.Scatter(
                x=_pc_labels, y=_cum_var, mode="lines+markers",
                name="Cumulative", line=dict(color=COLORS[1], width=2),
                marker=dict(size=6),
            ), row=1, col=3)
            _fig_pca.update_xaxes(title_text=_x_lbl,           row=1, col=1)
            _fig_pca.update_yaxes(title_text=_y_lbl,           row=1, col=1)
            _fig_pca.update_xaxes(title_text="Wavenumber (cm⁻¹)", row=1, col=2)
            _fig_pca.update_yaxes(title_text="Loading",            row=1, col=2)
            _fig_pca.update_yaxes(title_text="Variance (%)", range=[0, 105], row=1, col=3)
            _fig_pca.update_layout(height=380, legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(_fig_pca, use_container_width=True)
            st.caption(
                f"PCA of {_pr['scores'].shape[0]} spectra over {_pr['wn'][0]:.0f}–{_pr['wn'][-1]:.0f} cm⁻¹ "
                f"after selected preprocessing. "
                f"**a)** Score plot with {_x_lbl} on x and {_y_lbl} on y, coloured by {_c_lbl}. "
                f"**b)** PC loadings ({_n_pc} components) — peaks indicate spectral regions driving variance. "
                f"**c)** Individual (bars) and cumulative (line) explained variance per PC; "
                f"{_n_pc} PCs capture {sum(_var_exp):.1f}% of total spectral variance."
            )

        # ── b. Amide band decomposition ───────────────────────────────────
        st.divider()
        st.markdown("### b. Amide I band decomposition")
        _fig_caption(
            "Gaussians are fitted to each individual spectrum in the amide I region. "
            "Configure the preprocessing below, then click Run. "
            "The resulting Gaussian areas represent the fractional contribution of each secondary structure element."
        )
        _n_gauss = st.slider("Gaussian components", 2, 8, 6, key="n_gauss2")

        _items_amide = _pipeline_ui(
            "amide", ["spectral_cut", "spike_removal", "spectral_cut", "savgol", "area", "endpoint"],
            seed_defaults={"Spectral cut #2": {"wn_min": 1580, "wn_max": 1720, "use_gap": False}})

        # ── Gaussian centre constraints ───────────────────────────────────
        st.markdown("**Gaussian components**")
        # Presets from Fellows et al. 2020 (Applied Spectroscopy), Table 1
        # Tolerances = half the reported literature range
        _PRESETS = [
            {"label": "Aggregates / sidechains",    "centre": 1605, "tol": 8,  "fwhm_min": 15, "fwhm_max": 25},
            {"label": "β-sheet (inter. AP)",        "centre": 1618, "tol": 8,  "fwhm_min": 15, "fwhm_max": 25},
            {"label": "β-sheet (intra. P)",         "centre": 1635, "tol": 9,  "fwhm_min": 15, "fwhm_max": 25},
            {"label": "α-helix",                    "centre": 1653, "tol": 6,  "fwhm_min": 15, "fwhm_max": 25},
            {"label": "β-turn / random coil",       "centre": 1672, "tol": 13, "fwhm_min": 15, "fwhm_max": 25},
            {"label": "β-sheet (inter. AP) / s.c.", "centre": 1686, "tol": 11, "fwhm_min": 15, "fwhm_max": 25},
        ]
        _g_cens, _g_tols, _g_fwhm_min, _g_fwhm_max, _g_labels = [], [], [], [], []
        _g_cols = st.columns(_n_gauss)
        for _gi in range(_n_gauss):
            _pr_gi = _PRESETS[_gi] if _gi < len(_PRESETS) else {
                "label": f"Component {_gi+1}",
                "centre": 1605 + _gi * 20,
                "tol": 12, "fwhm_min": 15, "fwhm_max": 25,
            }
            with _g_cols[_gi]:
                _g_labels.append(st.text_input("Label", value=_pr_gi["label"], key=f"g{_gi}_label_b"))
                _g_cens.append(st.number_input("Centre (cm⁻¹)", value=_pr_gi["centre"], step=5,
                                                key=f"g{_gi}_cen_b"))
                _g_tols.append(st.number_input("± tol (cm⁻¹)", value=_pr_gi["tol"], min_value=1, step=1,
                                                key=f"g{_gi}_tol_b"))
                _g_fwhm_min.append(st.number_input("FWHM min (cm⁻¹)", value=_pr_gi["fwhm_min"], min_value=1, step=1,
                                                    key=f"g{_gi}_fwhm_min_b",
                                                    help="Lower bound on FWHM (σ = FWHM / 2.355)."))
                _g_fwhm_max.append(st.number_input("FWHM max (cm⁻¹)", value=_pr_gi["fwhm_max"], min_value=2, step=1,
                                                    key=f"g{_gi}_fwhm_max_b",
                                                    help="Upper bound on FWHM."))

        # ── PDB-informed amplitude priors ─────────────────────────────
        _pc1, _pc2, _pc3 = st.columns([1, 1.2, 3])
        _use_pdb = _pc1.toggle("PDB priors", value=False, key="use_pdb",
                               help="Use secondary structure content from PDB as amplitude initial guesses (p₀). "
                                    "Does not hard-constrain the fit — the optimiser can still deviate freely.")
        _pdb_id_in = _pc2.text_input("PDB ID", value="4INS", key="pdb_id_inp",
                                     label_visibility="collapsed", disabled=not _use_pdb)
        _amp_priors = [1.0 / _n_gauss] * _n_gauss  # flat default
        _pdb_fracs_display = None
        if _use_pdb and _pdb_id_in.strip():
            try:
                _h, _s, _o = _fetch_pdb_ss(_pdb_id_in.strip())
                _amp_priors = _pdb_amp_priors(_h, _s, _o, _n_gauss)
                _pdb_fracs_display = _amp_priors[:]
                _pc3.caption(
                    f"**{_pdb_id_in.upper()}** — α-helix: {_h*100:.0f}%, "
                    f"β-sheet: {_s*100:.0f}%, other: {_o*100:.0f}%"
                )
            except Exception as _pdb_err:
                _pc3.warning(f"PDB fetch failed: {_pdb_err}")

        _run_amide = st.button("▶ Run decomposition", key="btn_amide", on_click=_goto_further)

        if _run_amide:
            with st.spinner("Preprocessing spectra for amide decomposition…"):
                _X_am, _wn_am = _apply_pipeline(
                    _X_raw_all, _wn_raw_shared, _items_amide, "amide",
                )
            if _X_am.shape[1] < _n_gauss * 3:
                st.error("Too few wavenumber points for the requested number of Gaussians.")
            else:
                _mean_sp = _X_am.mean(axis=0)

                def _multi_gauss(x, *p):
                    y = np.zeros_like(x, dtype=float)
                    for _gi in range(len(p) // 3):
                        y += p[3*_gi] * np.exp(-((x - p[3*_gi+1])**2) / (2*p[3*_gi+2]**2))
                    return y

                _p0, _blo, _bhi = [], [], []
                for _gi in range(_n_gauss):
                    _cen      = float(_g_cens[_gi])
                    _tol      = float(_g_tols[_gi])
                    _smin     = float(_g_fwhm_min[_gi]) / 2.355
                    _smax     = float(_g_fwhm_max[_gi]) / 2.355
                    _p0  += [_amp_priors[_gi], _cen, max(_smin, min(10.0 / 2.355, _smax))]
                    _blo += [0,   max(float(_wn_am[0]),  _cen - _tol), _smin]
                    _bhi += [1.5, min(float(_wn_am[-1]), _cen + _tol), _smax]

                try:
                    # Fit mean spectrum (for display reference)
                    with st.spinner("Fitting mean spectrum…"):
                        _popt_mean, _ = _curve_fit(_multi_gauss, _wn_am, _mean_sp,
                                                    p0=_p0, bounds=(_blo, _bhi), maxfev=20000)
                    # Fit each individual spectrum; track successful indices for position lookup
                    _ind_params, _ind_ok, _ind_r2 = [], [], []
                    with st.spinner(f"Fitting {_X_am.shape[0]} individual spectra…"):
                        for _si in range(_X_am.shape[0]):
                            try:
                                _po, _ = _curve_fit(_multi_gauss, _wn_am, _X_am[_si],
                                                     p0=_popt_mean, bounds=(_blo, _bhi), maxfev=5000)
                                _ss_res = np.sum((_X_am[_si] - _multi_gauss(_wn_am, *_po)) ** 2)
                                _ss_tot = np.sum((_X_am[_si] - _X_am[_si].mean()) ** 2)
                                _r2 = float(1 - _ss_res / _ss_tot) if _ss_tot > 0 else 0.0
                                _ind_params.append(_po)
                                _ind_ok.append(_si)
                                _ind_r2.append(_r2)
                            except Exception:
                                pass

                    st.session_state["amide_result"] = dict(
                        wn=_wn_am, mean_sp=_mean_sp, popt=_popt_mean,
                        mc_params=np.array(_ind_params) if _ind_params else None,
                        ind_spectra=_X_am[_ind_ok] if _ind_ok else None,
                        ind_fnames=[_fname_list[_i] for _i in _ind_ok],
                        ind_dist=[_dist_list[_i] for _i in _ind_ok],
                        ind_ok=list(_ind_ok),
                        ind_r2=_ind_r2,
                        n_gauss=_n_gauss,
                        g_labels=list(_g_labels),
                        blo=_blo, bhi=_bhi, p0=_p0,
                        pdb_fracs=_pdb_fracs_display,
                    )
                except Exception as _exc:
                    st.error(f"Fitting failed: {_exc}")

        if "amide_result" in st.session_state:
            _ar       = st.session_state["amide_result"]
            _wn_am    = _ar["wn"]
            _ng       = _ar["n_gauss"]
            _mc       = _ar.get("mc_params")
            _ind_sp   = _ar.get("ind_spectra")
            _ind_fn   = _ar.get("ind_fnames", [])
            _ind_dt   = _ar.get("ind_dist", [])
            _ind_r2   = np.array(_ar.get("ind_r2", []))
            _glbls    = _ar.get("g_labels", [f"G{_i+1}" for _i in range(_ng)])
            _n_fitted = len(_mc) if _mc is not None else 0

            def _mg_plot(x, *p):
                y = np.zeros_like(x, dtype=float)
                for _gi in range(len(p) // 3):
                    y += p[3*_gi] * np.exp(-((x - p[3*_gi+1])**2) / (2*p[3*_gi+2]**2))
                return y

            # ── Filters ────────────────────────────────────────────────────
            # Derive concentrations from live _prot so unit changes are reflected
            _ind_ok_stored = _ar.get("ind_ok", [])
            if _prot is not None and _ind_ok_stored:
                _ind_prot = np.array([float(_prot[_i]) for _i in _ind_ok_stored])
            else:
                _ind_prot = np.array(_ar.get("ind_prot", [float("nan")] * _n_fitted))
            _r2_thresh = st.slider(
                "Min. R² (fit quality)",
                min_value=0.0, max_value=1.0, value=0.95, step=0.01,
                key="amide_r2_thresh",
                help="Exclude spectra whose Gaussian fit R² is below this threshold.",
            )
            _r2_mask   = (_ind_r2 >= _r2_thresh) if len(_ind_r2) == _n_fitted \
                         else np.ones(_n_fitted, dtype=bool)
            _valid_idx = np.where(_r2_mask)[0]
            _n_valid   = len(_valid_idx)

            # ── Slider: single-spectrum view ──────────────────────────────
            if _n_fitted > 0:
                if _n_valid == 0:
                    st.warning("No spectra pass the current R² threshold. Lower the threshold to view fits.")
                    _sel_pos, _sel = 0, 0
                elif _n_valid == 1:
                    st.info("1 spectrum passes the R² threshold.")
                    _sel_pos, _sel = 0, int(_valid_idx[0])
                else:
                    _sel_pos = st.slider("Spectrum #", 0, _n_valid - 1, 0, key="amide_slider")
                    _sel = int(_valid_idx[_sel_pos])
                _popt_sel = _mc[_sel]
                _fn  = _ind_fn[_sel]   if _sel < len(_ind_fn)   else "—"
                _dt  = _ind_dt[_sel]   if _sel < len(_ind_dt)   else float("nan")
                _cp  = _ind_prot[_sel] if _sel < len(_ind_prot) else float("nan")
                _r2v = float(_ind_r2[_sel]) if _sel < len(_ind_r2) else float("nan")
                _r2_ok = _r2v >= _r2_thresh
                _r2c   = ("green" if _r2v >= 0.95 else
                          ("darkorange" if _r2v >= _r2_thresh else "red"))
                _title_ind = (
                    f"Spectrum #{_sel} &nbsp;·&nbsp; {_fn}"
                    f" &nbsp;·&nbsp; {_dt:.2f} µm"
                    f" &nbsp;·&nbsp; {_cp:.2f} {unit}"
                    f" &nbsp;·&nbsp; <span style='color:{_r2c}'>R² = {_r2v:.4f}</span>"
                )
                # Compute average spectrum for overlay (reuse conc filter if available)
                _avg_for_overlay = None
                if _ind_sp is not None and len(_ind_sp) > 0:
                    if _global_conc_lo is not None:
                        _ov_idx = np.where(
                            (np.array(_ind_prot) >= _global_conc_lo) &
                            (np.array(_ind_prot) <= _global_conc_hi)
                        )[0]
                    else:
                        _ov_idx = np.arange(len(_ind_sp))
                    if len(_ov_idx) > 0:
                        _avg_for_overlay = _ind_sp[_ov_idx].mean(axis=0)

                _fig_single = go.Figure()
                # Average spectrum as grey reference line
                if _avg_for_overlay is not None:
                    _fig_single.add_trace(go.Scatter(
                        x=_wn_am, y=_avg_for_overlay, mode="lines",
                        name="Average spectrum",
                        line=dict(color="#aaaaaa", width=1.5, dash="dot"),
                    ))
                _fig_single.add_trace(go.Scatter(
                    x=_wn_am, y=_ind_sp[_sel], mode="lines",
                    name="This spectrum", line=dict(color="black", width=2),
                ))
                _fig_single.add_trace(go.Scatter(
                    x=_wn_am, y=_mg_plot(_wn_am, *_popt_sel), mode="lines",
                    name=f"Total fit (R²={_r2v:.3f})",
                    line=dict(color="red", dash="dash", width=2),
                ))
                _to_rgb = __import__("matplotlib.colors", fromlist=["to_rgb"]).to_rgb
                for _gi in range(_ng):
                    _a, _c, _s = _popt_sel[3*_gi], _popt_sel[3*_gi+1], _popt_sel[3*_gi+2]
                    _fwhm_i = 2.355 * _s
                    _col_i  = COLORS[_gi % len(COLORS)]
                    _fill_i = "rgba({},{},{},0.15)".format(
                        *[int(v*255) for v in _to_rgb(_col_i)])
                    _fig_single.add_trace(go.Scatter(
                        x=_wn_am,
                        y=_a * np.exp(-((_wn_am - _c)**2) / (2*_s**2)),
                        mode="lines", fill="tozeroy",
                        name=f"{_glbls[_gi]} ({_c:.0f} cm⁻¹, FWHM {_fwhm_i:.1f})",
                        line=dict(color=_col_i, width=1.5),
                        fillcolor=_fill_i,
                    ))
                _fig_single.update_layout(
                    title=dict(text=_title_ind, font=dict(size=12)),
                    xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Intensity (a.u.)",
                    yaxis=dict(tickformat="g"),
                    height=400, margin=dict(t=55),
                    legend=dict(orientation="v", x=1.02, xanchor="left"),
                )
                st.plotly_chart(_fig_single, use_container_width=True)
                st.caption(
                    f"Individual Gaussian decomposition of spectrum #{_sel} from {_fn}, "
                    f"acquired at {_dt:.2f} µm (protein: {_cp:.2f} {unit}). "
                    f"Black: this spectrum; grey dotted: average of {len(_ov_idx) if _avg_for_overlay is not None else 0} "
                    f"concentration-filtered spectra; red dashed: total Gaussian fit (R² = {_r2v:.4f}); "
                    f"coloured fills: individual components with their fitted centre positions. "
                    f"Amide I region: {_wn_am[0]:.0f}–{_wn_am[-1]:.0f} cm⁻¹."
                )
                if not _r2_ok:
                    st.warning(f"Poor fit: R² = {_r2v:.4f} < {_r2_thresh:.2f}")

            # ── Average spectrum fit (concentration-filtered) ────────────
            st.markdown("#### Average spectrum fit")
            if _ind_sp is not None and len(_ind_sp) > 0:
                def _multi_gauss_disp(x, *p):
                    y = np.zeros_like(x, dtype=float)
                    for _gi in range(len(p) // 3):
                        y += p[3*_gi] * np.exp(-((x - p[3*_gi+1])**2) / (2*p[3*_gi+2]**2))
                    return y

                if _global_conc_lo is not None:
                    _conc_idx = np.where(
                        (np.array(_ind_prot) >= _global_conc_lo) &
                        (np.array(_ind_prot) <= _global_conc_hi)
                    )[0]
                else:
                    _conc_idx = np.arange(len(_ind_sp))
                _n_conc   = len(_conc_idx)
                if _n_conc == 0:
                    st.warning("No spectra pass the current concentration range.")
                _sel_sp  = _ind_sp[_conc_idx] if _n_conc > 0 else _ind_sp
                _avg_sp  = _sel_sp.mean(axis=0)
                _std_sp  = _sel_sp.std(axis=0)
                _blo_ar = _ar.get("blo", [0]*(_ng*3))
                _bhi_ar = _ar.get("bhi", [np.inf]*(_ng*3))
                _p0_ar  = _ar.get("p0",  _ar["popt"].tolist())
                try:
                    from scipy.optimize import curve_fit as _cf2
                    _popt_avg, _ = _cf2(_multi_gauss_disp, _wn_am, _avg_sp,
                                        p0=_p0_ar, bounds=(_blo_ar, _bhi_ar), maxfev=20000)
                    _ss_res_avg = np.sum((_avg_sp - _multi_gauss_disp(_wn_am, *_popt_avg))**2)
                    _ss_tot_avg = np.sum((_avg_sp - _avg_sp.mean())**2)
                    _r2_avg = float(1 - _ss_res_avg / _ss_tot_avg) if _ss_tot_avg > 0 else 0.0

                    _fig_avg = go.Figure()
                    # Shaded ±1 SD region
                    _fig_avg.add_trace(go.Scatter(
                        x=np.concatenate([_wn_am, _wn_am[::-1]]),
                        y=np.concatenate([_avg_sp + _std_sp, (_avg_sp - _std_sp)[::-1]]),
                        fill="toself", fillcolor="rgba(128,128,128,0.25)",
                        line=dict(color="rgba(0,0,0,0)"), showlegend=True, name="± 1 SD",
                    ))
                    _fig_avg.add_trace(go.Scatter(
                        x=_wn_am, y=_avg_sp, name=f"Average ({_n_conc} spectra)",
                        line=dict(color="black", width=2)))
                    _y_total_avg = _multi_gauss_disp(_wn_am, *_popt_avg)
                    _fig_avg.add_trace(go.Scatter(
                        x=_wn_am, y=_y_total_avg, name=f"Total fit (R²={_r2_avg:.3f})",
                        line=dict(color="red", dash="dash", width=2)))
                    for _gi in range(_ng):
                        _amp_a = _popt_avg[3*_gi]
                        _cen_a = _popt_avg[3*_gi+1]
                        _sig_a = _popt_avg[3*_gi+2]
                        _y_gi  = _amp_a * np.exp(-((_wn_am - _cen_a)**2) / (2*_sig_a**2))
                        _lbl_a = _glbls[_gi] if _gi < len(_glbls) else f"G{_gi+1}"
                        _fwhm_a = 2.355 * _sig_a
                        _fig_avg.add_trace(go.Scatter(
                            x=_wn_am, y=_y_gi,
                            name=f"{_lbl_a} ({_cen_a:.0f} cm⁻¹, FWHM {_fwhm_a:.1f})",
                            line=dict(color=COLORS[_gi % len(COLORS)], width=1.5),
                            fill="tozeroy",
                            fillcolor=f"rgba({','.join(str(int(c*255)) for c in __import__('matplotlib.colors', fromlist=['to_rgb']).to_rgb(COLORS[_gi % len(COLORS)]))},0.15)",
                        ))
                    _fig_avg.update_layout(
                        xaxis_title="Wavenumber (cm⁻¹)", yaxis_title="Intensity (a.u.)",
                        yaxis=dict(tickformat="g"),
                        height=380, margin=dict(t=20),
                        legend=dict(orientation="v", x=1.02, xanchor="left"),
                    )
                    st.plotly_chart(_fig_avg, use_container_width=True)
                    _conc_filter_str = (
                        f"with PLS protein in [{_global_conc_lo:.2f}, {_global_conc_hi:.2f}] {unit}"
                        if _global_conc_lo is not None
                        else "(no concentration filter — build a PLS model to enable filtering)"
                    )
                    _fig_caption(
                        f"Gaussian decomposition of the mean amide I spectrum computed from {_n_conc} spectra "
                        f"{_conc_filter_str}. "
                        f"Grey shading: ±1 standard deviation across averaged spectra. "
                        f"Black: mean spectrum; red dashed: total Gaussian fit (R² = {_r2_avg:.4f}); "
                        f"coloured fills: individual components with fitted centre and FWHM shown in the legend. "
                        f"Component areas as fractions of the total fitted area are listed in the table below."
                    )
                    # Area percentages
                    _areas_avg = np.array([
                        _popt_avg[3*_gi] * _popt_avg[3*_gi+2] * np.sqrt(2 * np.pi)
                        for _gi in range(_ng)
                    ])
                    _total_area = _areas_avg.sum()
                    _pct_rows = []
                    for _gi in range(_ng):
                        _lbl_a = _glbls[_gi] if _gi < len(_glbls) else f"G{_gi+1}"
                        _row = {
                            "Component":       _lbl_a,
                            "Centre (cm⁻¹)":   f"{_popt_avg[3*_gi+1]:.1f}",
                            "FWHM (cm⁻¹)":    f"{2.355 * _popt_avg[3*_gi+2]:.1f}",
                            "Fitted area (%)": f"{100 * _areas_avg[_gi] / _total_area:.1f}",
                        }
                        if _pdb_fracs_display is not None and _gi < len(_pdb_fracs_display):
                            _row["PDB prior (%)"] = f"{100 * _pdb_fracs_display[_gi]:.1f}"
                        _pct_rows.append(_row)
                    st.dataframe(pd.DataFrame(_pct_rows), use_container_width=True, hide_index=True)
                except Exception as _exc_avg:
                    st.warning(f"Could not fit average spectrum: {_exc_avg}")
            else:
                st.info("No fitted spectra available. Run decomposition first.")

            # ── Summary table + per-component histograms (filtered) ───────
            _mc_f = _mc[_valid_idx] if (_mc is not None and _n_valid > 2) else None
            if _mc_f is not None:
                _rows = []
                for _gi in range(_ng):
                    _cts = _mc_f[:, 3*_gi+1]
                    _wds = _mc_f[:, 3*_gi+2]
                    _ars = _mc_f[:, 3*_gi] * _wds * np.sqrt(2 * np.pi)
                    _rows.append({
                        "Component":     _glbls[_gi] if _gi < len(_glbls) else f"G{_gi+1}",
                        "Centre (cm⁻¹)": f"{_cts.mean():.1f} ± {_cts.std():.1f}",
                        "FWHM (cm⁻¹)":  f"{(2.355*_wds).mean():.1f} ± {(2.355*_wds).std():.1f}",
                        "Area (a.u.)":   f"{_ars.mean():.3f} ± {_ars.std():.3f}",
                    })
                _n_poor = int(np.sum(_ind_r2 < _r2_thresh)) if len(_ind_r2) == _n_fitted else 0
                _fig_caption(
                    f"Summary statistics for {_n_valid} of {_n_fitted} fitted spectra "
                    f"passing the R² ≥ {_r2_thresh:.2f} quality threshold "
                    f"({_n_poor} spectra excluded). "
                    f"Centre and FWHM values show mean ± SD; areas are in absolute units. "
                    f"Histograms below show the distribution of fitted peak centres across all valid spectra."
                )
                st.dataframe(pd.DataFrame(_rows), use_container_width=True, hide_index=True)

                _hist_titles = [
                    f"{chr(ord('a') + _gi)})  {_glbls[_gi] if _gi < len(_glbls) else f'G{_gi+1}'}"
                    for _gi in range(_ng)
                ]
                _fig_hist = make_subplots(
                    rows=1, cols=_ng,
                    subplot_titles=_hist_titles,
                )
                for _gi in range(_ng):
                    _fig_hist.add_trace(go.Histogram(
                        x=_mc_f[:, 3*_gi+1], name=_glbls[_gi] if _gi < len(_glbls) else f"G{_gi+1}",
                        marker_color=COLORS[_gi % len(COLORS)], opacity=0.7,
                        showlegend=False,
                    ), row=1, col=_gi+1)
                    _fig_hist.update_xaxes(title_text="Peak centre (cm⁻¹)", row=1, col=_gi+1)
                    _fig_hist.update_yaxes(title_text="Count" if _gi == 0 else "", row=1, col=_gi+1)
                _fig_hist.update_layout(height=260, margin=dict(t=40))
                st.plotly_chart(_fig_hist, use_container_width=True)
                st.caption(
                    f"Histograms of fitted peak centre positions for each Gaussian component "
                    f"({_ng} components, {_wn_am[0]:.0f}–{_wn_am[-1]:.0f} cm⁻¹) "
                    f"across {_n_valid} spectra with R² ≥ {_r2_thresh:.2f}. "
                    f"Narrow, symmetric distributions indicate stable peak positions consistent with a single secondary structure element; "
                    f"broad or skewed distributions suggest heterogeneity or model instability for that component."
                )

        # ── c. Peak ratio ─────────────────────────────────────────────────
        st.divider()
        st.markdown("### c. Peak ratio comparison")
        _fig_caption(
            "Integrates spectral intensity in two user-defined wavenumber windows and plots "
            "their ratio along the linescan position and against any PLS or MCR score. "
            "A simple, model-free way to track how the relative intensity of two spectral bands "
            "changes across the sample."
        )
        _cc1, _cc2, _cc3 = st.columns(3)
        _pk1 = _cc1.number_input("Peak 1 centre (cm⁻¹)", value=1003, step=5, key="pk1",
                                   help="Centre of the first integration window.")
        _pk2 = _cc2.number_input("Peak 2 centre (cm⁻¹)", value=1655, step=5, key="pk2",
                                   help="Centre of the second integration window.")
        _hw  = _cc3.number_input("Half-window (cm⁻¹)", value=15, min_value=1,
                                   step=5, key="pk_hw",
                                   help="Intensity summed over [centre ± half-window]")
        _pr_lbl = st.selectbox("Plot ratio vs", list(_label_opts.keys()), key="pr_label")
        _items_peak = _pipeline_ui(
            "peak", ["spectral_cut", "spike_removal", "rubberband"])
        _run_peak = st.button("▶ Compute ratio", key="btn_peak", on_click=_goto_further)

        if _run_peak:
            with st.spinner("Preprocessing spectra for peak ratio…"):
                _X_peak, _wn_peak = _apply_pipeline(
                    _X_raw_all, _wn_raw_shared, _items_peak, "peak"
                )

            def _integrate(X, wn, centre, hw):
                _m = (wn >= centre - hw) & (wn <= centre + hw)
                return X[:, _m].sum(axis=1) if _m.any() else np.zeros(X.shape[0])

            _a1    = _integrate(_X_peak, _wn_peak, _pk1, _hw)
            _a2    = _integrate(_X_peak, _wn_peak, _pk2, _hw)
            _ratio = np.where(_a2 > 0, _a1 / _a2, np.nan)
            st.session_state["ratio_result"] = dict(
                ratio=_ratio,
                fnames=_fname_list,
                distances=_dist_list,
                pk1=_pk1, pk2=_pk2, hw=_hw,
            )

        if "ratio_result" in st.session_state:
            _rr     = st.session_state["ratio_result"]
            _ratio  = _rr["ratio"]
            _pr_col = _label_opts[_pr_lbl]

            _fig_pr = make_subplots(
                rows=1, cols=2,
                subplot_titles=[
                    "a)  Ratio profile along linescan",
                    f"b)  Ratio vs {_pr_lbl}",
                ],
            )
            _offset = 0
            for _fi, (_fn, _r) in enumerate(results_all.items()):
                _n_r   = _r["X_proc"].shape[0]
                _fig_pr.add_trace(go.Scatter(
                    x=_r["distance"], y=_ratio[_offset:_offset + _n_r],
                    mode="lines", name=_fn,
                    line=dict(color=COLORS[_fi % len(COLORS)], width=1.5),
                ), row=1, col=1)
                _offset += _n_r

            _fig_pr.add_trace(go.Scatter(
                x=_pr_col, y=_ratio, mode="markers",
                marker=dict(color=COLORS[0], size=5, opacity=0.5,
                            line=dict(color="black", width=0.3)),
                showlegend=False,
            ), row=1, col=2)
            _yr = f"I({_rr['pk1']}) / I({_rr['pk2']})"
            _fig_pr.update_xaxes(title_text=dist_label, row=1, col=1)
            _fig_pr.update_xaxes(title_text=_pr_lbl,    row=1, col=2)
            _fig_pr.update_yaxes(title_text=_yr,         row=1, col=1)
            _fig_pr.update_yaxes(title_text=_yr,         row=1, col=2)
            _fig_pr.update_layout(height=360, legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(_fig_pr, use_container_width=True)
            st.caption(
                f"Band intensity ratio I({_rr['pk1']} cm⁻¹) / I({_rr['pk2']} cm⁻¹) "
                f"integrated over ±{_rr['hw']} cm⁻¹ windows, computed from {len(_ratio)} spectra "
                f"across {len(results_all)} linescan(s). "
                f"**a)** Ratio as a function of position along the linescan — "
                f"spatial variation reflects local changes in the relative intensities of the two bands. "
                f"**b)** Ratio plotted against {_pr_lbl} — "
                f"a systematic trend here indicates that the two bands change in proportion to each other."
            )


# ── Image overlay ─────────────────────────────────────────────────────────────
with tab_image:
    st.subheader("🗺️ Linescan score overlay on microscopy image")
    st.caption(
        "Upload a brightfield/fluorescence image of the sample, provide its physical dimensions "
        "and stage position, then choose a score to visualise as a colour-coded overlay along the linescan path."
    )

    if "results" not in st.session_state:
        st.info("Run the analysis first to generate scores for overlay.")
    else:
        _ov_results = st.session_state["results"]
        _ov_unit    = st.session_state.get("unit", "mg/mL")

        # ── Magnification preset ───────────────────────────────────────────
        _MAG_PRESETS = {
            "100×  (115.72 × 86.79 µm)": dict(
                w=115.72, h=86.79, xy_res=0.306, z_res=0.930,
            ),
            "20×  (553.224 × 414.918 µm)": dict(
                w=553.224, h=414.918, xy_res=0.550, z_res=3.06,
            ),
            "5×  (2209.168 × 1656.876 µm)": dict(
                w=2209.168, h=1656.876, xy_res=2.12, z_res=44.8,
            ),
            "Custom": None,
        }
        _mag_choice = st.selectbox(
            "Objective magnification",
            list(_MAG_PRESETS.keys()),
            key="ov_mag",
            help="Select a preset to auto-fill image dimensions and spatial resolution, "
                 "or choose Custom to enter values manually.",
        )
        _preset = _MAG_PRESETS[_mag_choice]

        # ── Image upload ──────────────────────────────────────────────────
        st.markdown("#### Image")
        _img_col1, _img_col2 = st.columns([1, 1])

        with _img_col1:
            _img_file = st.file_uploader(
                "Microscopy image (.bmp / .png / .tif)",
                type=["bmp", "png", "tif", "tiff", "jpg", "jpeg"],
                key="overlay_img",
            )
            _pos_file = st.file_uploader(
                "Position file (.txt) — optional",
                type=["txt"],
                key="overlay_pos",
                help="WITec position text file containing 'Position X' and 'Position Y' lines. "
                     "If not provided, set the image centre position manually below.",
            )

        with _img_col2:
            st.markdown("**Physical image dimensions**")
            _dim_c1, _dim_c2 = st.columns(2)
            _default_w = _preset["w"] if _preset else 115.72
            _default_h = _preset["h"] if _preset else 86.79
            _img_w_um = _dim_c1.number_input(
                "Width (µm)", value=_default_w, min_value=0.1, step=1.0, key="ov_img_w_um",
            )
            _img_h_um = _dim_c2.number_input(
                "Height (µm)", value=_default_h, min_value=0.1, step=1.0, key="ov_img_h_um",
            )

            st.markdown("**Spatial resolution**")
            _res_c1, _res_c2 = st.columns(2)
            _default_xy_res = _preset["xy_res"] if _preset else 0.306
            _default_z_res  = _preset["z_res"]  if _preset else 0.930
            _xy_res = _res_c1.number_input(
                "XY resolution (µm)", value=_default_xy_res, min_value=0.001,
                format="%.3f", step=0.01, key="ov_xy_res",
                help="Spatial resolution in the XY plane — used to set the default dot size.",
            )
            _z_res  = _res_c2.number_input(
                "Z resolution (µm)",  value=_default_z_res,  min_value=0.001,
                format="%.3f", step=0.01, key="ov_z_res",
                help="Depth resolution — informational.",
            )

            st.markdown("**Image centre position (µm)** — overridden by position file if uploaded")
            _pos_c1, _pos_c2 = st.columns(2)
            _manual_cx = _pos_c1.number_input("Centre X (µm)", value=0.0, step=1.0, key="ov_cx")
            _manual_cy = _pos_c2.number_input("Centre Y (µm)", value=0.0, step=1.0, key="ov_cy")

        # ── Score selection ───────────────────────────────────────────────
        st.markdown("#### Score to overlay")

        _ov_score_opts = {}
        _ov_r0 = _ov_results[list(_ov_results.keys())[0]]
        if "pls_protein" in _ov_r0 and _ov_r0["pls_protein"] is not None:
            _ov_score_opts[f"PLS Protein ({_ov_unit})"] = "pls_protein"
        if "pls_peg" in _ov_r0 and _ov_r0["pls_peg"] is not None:
            _ov_score_opts["PLS Molecular crowder (wt%)"] = "pls_peg"
        if "pls_salt" in _ov_r0 and _ov_r0["pls_salt"] is not None:
            _ov_score_opts["PLS Salt (mM)"] = "pls_salt"
        if "C_mcr" in _ov_r0 and _ov_r0["C_mcr"] is not None:
            _n_mcr = _ov_r0["C_mcr"].shape[1]
            for _ci in range(_n_mcr):
                _ov_score_opts[f"MCR component {_ci+1}"] = ("C_mcr", _ci)

        _ov_col1, _ov_col2 = st.columns([2, 1])
        _ov_score_label = _ov_col1.selectbox(
            "Score", list(_ov_score_opts.keys()), key="ov_score_label",
        ) if _ov_score_opts else None
        _ov_cmap = _ov_col2.selectbox(
            "Colour map", ["Viridis", "Plasma", "RdBu_r", "Turbo", "Inferno"],
            key="ov_cmap",
        )

        _ov_scan_select = st.selectbox(
            "Linescan to show", list(_ov_results.keys()), key="ov_scan_sel",
        )

        # ── Crop controls ─────────────────────────────────────────────────
        st.markdown("#### Image crop (µm from each edge, 0 = no crop)")
        _crop_c1, _crop_c2, _crop_c3, _crop_c4 = st.columns(4)
        _crop_left   = _crop_c1.number_input("Left (µm)",   value=0.0, min_value=0.0, step=1.0, key="ov_crop_left")
        _crop_right  = _crop_c2.number_input("Right (µm)",  value=0.0, min_value=0.0, step=1.0, key="ov_crop_right")
        _crop_top    = _crop_c3.number_input("Top (µm)",    value=0.0, min_value=0.0, step=1.0, key="ov_crop_top")
        _crop_bottom = _crop_c4.number_input("Bottom (µm)", value=0.0, min_value=0.0, step=1.0, key="ov_crop_bottom")

        # ── Render overlay ────────────────────────────────────────────────
        if _img_file is not None and _ov_score_label is not None:
            import plotly.express as _px_ov
            from PIL import Image as _PIL_Image

            # Detect scan mode
            _ov_sm = st.session_state.get("scan_mode", "xy")

            # Load image
            _pil_img = _PIL_Image.open(_img_file).convert("RGB")
            _img_arr = np.array(_pil_img)
            _img_w_px, _img_h_px = _pil_img.size  # (width, height) in pixels

            # Apply crop: convert µm to pixels and slice
            _px_per_um_x = _img_w_px / _img_w_um
            _px_per_um_y = _img_h_px / _img_h_um
            _crop_l_px = int(round(_crop_left   * _px_per_um_x))
            _crop_r_px = int(round(_crop_right  * _px_per_um_x))
            _crop_t_px = int(round(_crop_top    * _px_per_um_y))
            _crop_b_px = int(round(_crop_bottom * _px_per_um_y))
            _crop_r_px = _img_w_px - _crop_r_px if _crop_r_px > 0 else _img_w_px
            _crop_b_px = _img_h_px - _crop_b_px if _crop_b_px > 0 else _img_h_px
            _img_arr   = _img_arr[_crop_t_px:_crop_b_px, _crop_l_px:_crop_r_px]
            _img_w_px_c = _img_arr.shape[1]
            _img_h_px_c = _img_arr.shape[0]

            # Auto dot size: diameter of one resolution element in image pixels,
            # scaled to display pixels (~800px display width assumed).
            _display_w_px = 800
            _auto_dot = max(4, round(_xy_res * _display_w_px / _img_w_um))
            _ov_dot_size = st.slider(
                "Dot size (px)",
                min_value=1, max_value=40, value=_auto_dot, key="ov_dot_size",
                help=f"Auto-calculated from XY resolution ({_xy_res:.3f} µm) relative to "
                     f"field width ({_img_w_um:.2f} µm). Adjust if needed.",
            )

            # Parse position file if provided
            _img_cx_um = _manual_cx
            _img_cy_um = _manual_cy
            if _pos_file is not None:
                _pos_text = _pos_file.read().decode("utf-8", errors="replace")
                _pos_file.seek(0)
                for _pline in _pos_text.splitlines():
                    if "Position X" in _pline and ":" in _pline:
                        try: _img_cx_um = float(_pline.split(":")[1].strip().split()[0])
                        except: pass
                    elif "Position Y" in _pline and ":" in _pline:
                        try: _img_cy_um = float(_pline.split(":")[1].strip().split()[0])
                        except: pass

            # Get scan result and extract scores / distance
            _ov_r    = _ov_results[_ov_scan_select]
            _ov_dist = np.asarray(_ov_r["distance"], dtype=float)
            _ov_key  = _ov_score_opts[_ov_score_label]
            if isinstance(_ov_key, tuple):
                _ov_scores = _ov_r[_ov_key[0]][:, _ov_key[1]]
            else:
                _ov_scores = _ov_r[_ov_key]
            _ov_scores = np.asarray(_ov_scores, dtype=float)

            # Stage-coordinate → cropped-image-pixel helpers
            # Image top-left corner in stage µm coords:
            _img_left_um = _img_cx_um - _img_w_um / 2
            _img_top_um  = _img_cy_um - _img_h_um / 2

            _flip_y = st.checkbox(
                "Flip Y axis", value=False, key="ov_flip_y",
                help="Tick if dots appear vertically mirrored. "
                     "Some instruments report stage Y increasing upward while image Y increases downward.",
            )

            def _stage_to_px(xs, ys):
                """Convert stage µm arrays to cropped-image pixel arrays."""
                xp = (np.asarray(xs, dtype=float) - _img_left_um) * _px_per_um_x - _crop_l_px
                if _flip_y:
                    yp = (_img_cy_um + _img_h_um / 2 - np.asarray(ys, dtype=float)) * _px_per_um_y - _crop_t_px
                else:
                    yp = (np.asarray(ys, dtype=float) - _img_top_um) * _px_per_um_y - _crop_t_px
                return xp, yp

            # ── Z-scan: image (with dot + scale bar) + score-vs-depth ────
            if _ov_sm == "z":
                # Dot at the mean XY stage position of the z-scan
                _pos_x = _ov_r.get("positions", {}).get("x", [_img_cx_um])
                _pos_y = _ov_r.get("positions", {}).get("y", [_img_cy_um])
                _scan_x_px, _scan_y_px = _stage_to_px(
                    [float(np.mean(_pos_x))], [float(np.mean(_pos_y))]
                )
                _scan_x_px = float(_scan_x_px[0])
                _scan_y_px = float(_scan_y_px[0])

                # Mean score for the single dot colour
                _dot_color_mean = float(np.nanmean(_ov_scores))

                _fig_img = _px_ov.imshow(_img_arr, binary_backend="jpg")

                # Dot at the scan XY position, coloured by mean score
                _fig_img.add_trace(go.Scatter(
                    x=[_scan_x_px],
                    y=[_scan_y_px],
                    mode="markers",
                    marker=dict(
                        color=[_dot_color_mean],
                        colorscale=_ov_cmap,
                        cmin=float(_ov_scores.min()),
                        cmax=float(_ov_scores.max()),
                        size=max(10, _ov_dot_size * 2),
                        symbol="circle-open",
                        line=dict(color="white", width=2),
                        showscale=False,
                    ),
                    hovertemplate=(
                        f"Z-scan position<br>"
                        f"Mean {_ov_score_label}: {_dot_color_mean:.3f}<extra></extra>"
                    ),
                    name="Z-scan position",
                ))

                # Scale bar (physical, using image physical width)
                _sb_um   = max(1.0, round(_img_w_um * 0.15 / 5) * 5)  # ~15 % of width, rounded to 5 µm
                _sb_px   = _sb_um * (_img_w_px_c / _img_w_um)
                _sb_x0   = 0.05 * _img_w_px_c
                _sb_y0   = 0.93 * _img_h_px_c
                # White semi-transparent background box behind the label + bar
                _sb_pad  = 0.015 * _img_h_px_c
                _fig_img.add_shape(type="rect",
                    x0=_sb_x0 - _sb_pad, y0=_sb_y0 - 0.06 * _img_h_px_c,
                    x1=_sb_x0 + _sb_px + _sb_pad, y1=_sb_y0 + _sb_pad,
                    fillcolor="rgba(255,255,255,0.5)", line=dict(width=0),
                )
                _fig_img.add_shape(type="line",
                    x0=_sb_x0, y0=_sb_y0, x1=_sb_x0 + _sb_px, y1=_sb_y0,
                    line=dict(color="black", width=2),
                )
                _fig_img.add_annotation(
                    x=_sb_x0 + _sb_px / 2, y=_sb_y0 - 0.035 * _img_h_px_c,
                    text=f"{_sb_um:.0f} µm", showarrow=False,
                    font=dict(color="black", size=12),
                )

                # Frame around the image
                _fig_img.add_shape(type="rect",
                    x0=0, y0=0, x1=_img_w_px_c, y1=_img_h_px_c,
                    line=dict(color="white", width=2),
                    fillcolor="rgba(0,0,0,0)",
                )

                _fig_img.update_layout(
                    height=500,
                    margin=dict(l=0, r=0, t=30, b=0),
                    xaxis=dict(visible=False),
                    yaxis=dict(visible=False, scaleanchor="x"),
                    plot_bgcolor="black",
                    showlegend=False,
                    title=dict(text=f"Image — {_ov_scan_select}", font=dict(size=13)),
                )

                # Depth plot (half the display width of the image column → [2, 1] ratio)
                _z_col1, _z_col2 = st.columns([2, 1])

                with _z_col1:
                    st.plotly_chart(_fig_img, use_container_width=True)

                with _z_col2:
                    _fig_depth = go.Figure()
                    _fig_depth.add_trace(go.Scatter(
                        x=_ov_scores,
                        y=_ov_dist,
                        mode="markers",
                        marker=dict(
                            color=_ov_scores,
                            colorscale=_ov_cmap,
                            size=_ov_dot_size,
                            line=dict(color="black", width=0.5),
                            colorbar=dict(
                                title=dict(text=_ov_score_label, side="right"),
                                thickness=12,
                                len=1.0,
                                y=0.5,
                                yanchor="middle",
                                x=1.01,
                                xanchor="left",
                            ),
                            showscale=True,
                        ),
                        hovertemplate=(
                            f"<b>{_ov_score_label}</b>: %{{x:.3f}}<br>"
                            "Depth: %{y:.2f} µm<extra></extra>"
                        ),
                        name=_ov_score_label,
                    ))
                    _fig_depth.update_layout(
                        height=500,
                        margin=dict(l=10, r=20, t=30, b=40),
                        xaxis=dict(title=_ov_score_label),
                        yaxis=dict(title="Depth (µm)"),
                        title=dict(text=f"{_ov_score_label} vs Depth", font=dict(size=13)),
                    )
                    st.plotly_chart(_fig_depth, use_container_width=True)

                _fig_ov = _fig_depth  # for download

            # ── XY-scan: true 2D overlay using stage coordinates ──────────
            else:
                _pos_x = np.asarray(_ov_r.get("positions", {}).get("x", []), dtype=float)
                _pos_y = np.asarray(_ov_r.get("positions", {}).get("y", []), dtype=float)
                _x_px_2d, _y_px_2d = _stage_to_px(_pos_x, _pos_y)

                _fig_ov = _px_ov.imshow(_img_arr, binary_backend="jpg")
                _fig_ov.add_trace(go.Scatter(
                    x=_x_px_2d,
                    y=_y_px_2d,
                    mode="markers",
                    marker=dict(
                        color=_ov_scores,
                        colorscale=_ov_cmap,
                        size=_ov_dot_size,
                        line=dict(color="black", width=0.5),
                        colorbar=dict(
                            title=dict(text=_ov_score_label, side="right"),
                            thickness=15,
                            len=1.0,
                            y=0.5,
                            yanchor="middle",
                            x=1.01,
                            xanchor="left",
                        ),
                        showscale=True,
                    ),
                    hovertemplate=(
                        f"<b>{_ov_score_label}</b>: %{{marker.color:.3f}}<br>"
                        "Distance: %{customdata:.2f} µm<extra></extra>"
                    ),
                    customdata=_ov_dist,
                    name=_ov_score_label,
                ))

                # Scale bar (physical, based on image dimensions)
                _bar_um = max(1.0, round(_img_w_um * 0.15 / 5) * 5)
                _bar_px = _bar_um * (_img_w_px_c / _img_w_um)
                _bar_x0 = 0.05 * _img_w_px_c
                _bar_y0 = 0.93 * _img_h_px_c
                _sb_pad = 0.015 * _img_h_px_c
                _fig_ov.add_shape(type="rect",
                    x0=_bar_x0 - _sb_pad, y0=_bar_y0 - 0.06 * _img_h_px_c,
                    x1=_bar_x0 + _bar_px + _sb_pad, y1=_bar_y0 + _sb_pad,
                    fillcolor="rgba(255,255,255,0.5)", line=dict(width=0),
                )
                _fig_ov.add_shape(type="line",
                    x0=_bar_x0, y0=_bar_y0, x1=_bar_x0 + _bar_px, y1=_bar_y0,
                    line=dict(color="black", width=2),
                )
                _fig_ov.add_annotation(
                    x=_bar_x0 + _bar_px / 2, y=_bar_y0 - 0.035 * _img_h_px_c,
                    text=f"{_bar_um:.0f} µm", showarrow=False,
                    font=dict(color="black", size=12),
                )
                # Frame
                _fig_ov.add_shape(type="rect",
                    x0=0, y0=0, x1=_img_w_px_c, y1=_img_h_px_c,
                    line=dict(color="white", width=2), fillcolor="rgba(0,0,0,0)",
                )

                _fig_ov.update_layout(
                    height=600,
                    margin=dict(l=0, r=0, t=30, b=0),
                    xaxis=dict(visible=False),
                    yaxis=dict(visible=False, scaleanchor="x"),
                    plot_bgcolor="black",
                    title=dict(
                        text=f"{_ov_scan_select} — {_ov_score_label}",
                        font=dict(size=13),
                    ),
                )
                st.plotly_chart(_fig_ov, use_container_width=True)
                st.caption(
                    f"Each dot is placed at its stage XY position relative to the image centre "
                    f"({_img_cx_um:.1f}, {_img_cy_um:.1f}) µm. "
                    f"Image physical size: {_img_w_um:.1f} × {_img_h_um:.1f} µm."
                )

            # Store overlay for download
            st.session_state["overlay_fig_html"] = _fig_ov.to_html(include_plotlyjs="cdn", full_html=True)
            st.session_state["overlay_fig_label"] = _ov_score_label
            st.session_state["overlay_fig_scan"]  = _ov_scan_select
            try:
                import plotly.io as _pio
                st.session_state["overlay_fig_pdf"] = _pio.to_image(_fig_ov, format="pdf")
            except Exception:
                st.session_state["overlay_fig_pdf"] = None

        elif _img_file is None:
            st.info("Upload a microscopy image above to display the overlay.")
        elif not _ov_score_opts:
            st.warning("No scores available — run the analysis first.")


# ── Download ──────────────────────────────────────────────────────────────────
with tab_download:
    if "results" not in st.session_state or not st.session_state["results"]:
        st.info("Run the analysis first.")
    else:
        results_all = st.session_state["results"]
        unit        = st.session_state.get("unit", "mg/mL")
        sm          = st.session_state.get("scan_mode", "xy")

        # ── PLS / MCR linescan results ──────────────────────────────────────
        st.subheader("PLS / MCR linescan results")

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, res in results_all.items():
                safe  = os.path.splitext(fname)[0].replace(" ", "_")
                xlsx  = an.results_to_excel_bytes(res, sm, unit, "conc")
                zf.writestr(f"{safe}_results.xlsx", xlsx)

        st.download_button(
            "⬇  Download all linescan results (ZIP)",
            data=zip_buf.getvalue(),
            file_name="de_condensate_results.zip",
            mime="application/zip",
            use_container_width=True,
        )

        with st.expander("Individual linescan files"):
            for fname, res in results_all.items():
                safe = os.path.splitext(fname)[0].replace(" ", "_")
                xlsx = an.results_to_excel_bytes(res, sm, unit, "conc")
                st.download_button(
                    f"⬇  {safe}_results.xlsx",
                    data=xlsx,
                    file_name=f"{safe}_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{fname}",
                )

        # ── Further analysis results ────────────────────────────────────────
        st.divider()
        st.subheader("Further analysis results")

        _has_pca   = "pca_result"   in st.session_state
        _has_amide = "amide_result" in st.session_state
        _has_ratio = "ratio_result" in st.session_state

        if not any([_has_pca, _has_amide, _has_ratio]):
            st.info("Run analyses in the Further analysis tab to enable downloads here.")
        else:
            import openpyxl as _opxl

            if _has_pca:
                _pr = st.session_state["pca_result"]
                _pca_buf = io.BytesIO()
                _wb = _opxl.Workbook()

                # Scores sheet
                _ws = _wb.active
                _ws.title = "PCA scores"
                _n_pc = _pr["scores"].shape[1]
                _var  = _pr["var_exp"]
                _ws.append([f"PC{_i+1} ({_var[_i]:.2f}%)" for _i in range(_n_pc)])
                for row in _pr["scores"].tolist():
                    _ws.append(row)

                # Loadings sheet
                _ws2 = _wb.create_sheet("PCA loadings")
                _ws2.append(["Wavenumber"] + [f"PC{_i+1}" for _i in range(_n_pc)])
                for _wni, _wn_val in enumerate(_pr["wn"]):
                    _ws2.append([float(_wn_val)] + [float(_pr["loadings"][_k, _wni]) for _k in range(_n_pc)])

                # Variance sheet
                _ws3 = _wb.create_sheet("Explained variance")
                _ws3.append(["PC", "Individual (%)", "Cumulative (%)"])
                _cum = 0.0
                for _i, _v in enumerate(_var):
                    _cum += _v
                    _ws3.append([f"PC{_i+1}", round(float(_v), 4), round(_cum, 4)])

                _wb.save(_pca_buf)
                st.download_button(
                    "⬇  PCA results (.xlsx)",
                    data=_pca_buf.getvalue(),
                    file_name="further_pca.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_pca",
                )

            if _has_amide:
                _ar = st.session_state["amide_result"]
                _amide_buf = io.BytesIO()
                _wb2 = _opxl.Workbook()

                # Per-spectrum Gaussian parameters
                _ws4 = _wb2.active
                _ws4.title = "Gaussian fits"
                _ng  = _ar["n_gauss"]
                _glbls = _ar.get("g_labels", [f"G{_k+1}" for _k in range(_ng)])
                _hdr = ["file", "distance", "R²"]
                for _gl in _glbls:
                    _hdr += [f"{_gl}_amplitude", f"{_gl}_centre", f"{_gl}_sigma"]
                _ws4.append(_hdr)
                _mc = _ar.get("mc_params")
                if _mc is not None:
                    for _i, _row in enumerate(_mc):
                        _fn  = _ar["ind_fnames"][_i] if _i < len(_ar["ind_fnames"]) else ""
                        _dt  = _ar["ind_dist"][_i]   if _i < len(_ar["ind_dist"])   else ""
                        _r2  = float(_ar["ind_r2"][_i]) if _i < len(_ar["ind_r2"])  else ""
                        _ws4.append([_fn, _dt, _r2] + [float(v) for v in _row])

                # Mean spectrum + fit
                _ws5 = _wb2.create_sheet("Mean spectrum")
                _ws5.append(["Wavenumber", "Mean intensity", "Fitted"])
                _fit_y = [None] * len(_ar["wn"])
                if _ar.get("popt") is not None:
                    try:
                        _p = _ar["popt"]
                        _fit_y = np.zeros(len(_ar["wn"]))
                        for _gi in range(len(_p) // 3):
                            _A, _mu, _sig = _p[3*_gi], _p[3*_gi+1], _p[3*_gi+2]
                            _fit_y += _A * np.exp(-0.5 * ((_ar["wn"] - _mu) / _sig) ** 2)
                    except Exception:
                        _fit_y = [None] * len(_ar["wn"])
                for _wi, _wv in enumerate(_ar["wn"]):
                    _fv = _fit_y[_wi] if _fit_y[_wi] is not None else ""
                    _ws5.append([float(_wv), float(_ar["mean_sp"][_wi]),
                                 float(_fv) if _fv != "" else ""])

                _wb2.save(_amide_buf)
                st.download_button(
                    "⬇  Amide decomposition results (.xlsx)",
                    data=_amide_buf.getvalue(),
                    file_name="further_amide.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_amide",
                )

            if _has_ratio:
                _rr = st.session_state["ratio_result"]
                _ratio_buf = io.BytesIO()
                _wb3 = _opxl.Workbook()
                _ws6 = _wb3.active
                _ws6.title = f"I({_rr['pk1']}) over I({_rr['pk2']})"
                _ws6.append(["file", "distance (µm)", f"I({_rr['pk1']})/I({_rr['pk2']})"])
                for _i, (_fn, _dt) in enumerate(zip(_rr["fnames"], _rr["distances"])):
                    _ws6.append([_fn, float(_dt), float(_rr["ratio"][_i]) if not np.isnan(_rr["ratio"][_i]) else ""])
                _wb3.save(_ratio_buf)
                st.download_button(
                    f"⬇  Peak ratio I({_rr['pk1']})/I({_rr['pk2']}) (.xlsx)",
                    data=_ratio_buf.getvalue(),
                    file_name="further_peak_ratio.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_ratio",
                )

        # ── Image overlay ───────────────────────────────────────────────────
        st.divider()
        st.subheader("Image overlay")
        if "overlay_fig_html" not in st.session_state:
            st.info("Generate an overlay in the Image overlay tab to enable download here.")
        else:
            _ov_label = st.session_state.get("overlay_fig_label", "score")
            _ov_scan  = st.session_state.get("overlay_fig_scan",  "linescan")
            _ov_safe  = _ov_scan.replace(" ", "_").replace(".", "_")
            _dl_c1, _dl_c2 = st.columns(2)
            _dl_c1.download_button(
                f"⬇  Download overlay (.html)",
                data=st.session_state["overlay_fig_html"].encode("utf-8"),
                file_name=f"overlay_{_ov_safe}.html",
                mime="text/html",
                key="dl_overlay_html",
                use_container_width=True,
            )
            _pdf_bytes = st.session_state.get("overlay_fig_pdf")
            if _pdf_bytes:
                _dl_c2.download_button(
                    f"⬇  Download overlay (.pdf)",
                    data=_pdf_bytes,
                    file_name=f"overlay_{_ov_safe}.pdf",
                    mime="application/pdf",
                    key="dl_overlay_pdf",
                    use_container_width=True,
                )
            else:
                _dl_c2.info("PDF export requires the `kaleido` package (`pip install kaleido`).")
            st.caption(
                f"**{_ov_scan} — {_ov_label}**. "
                "HTML: interactive, zoomable in any browser. "
                "PDF: static vector graphic, suitable for publications."
            )


# ── Tutorial ───────────────────────────────────────────────────────────────────
with tab_tutorial:
    st.subheader("Tutorial")

    st.markdown(
        "**De-condensate** is an interactive tool for spatially resolved, quantitative "
        "spectral analysis of Raman linescans recorded across protein condensates. "
        "All analytical steps are individually selectable — you can run only the parts "
        "relevant to your experiment."
    )

    # ── What you can do ───────────────────────────────────────────────────────
    with st.expander("What you can do — and what is optional", expanded=True):
        st.markdown(
            "Every analysis step is independent. Mix and match as needed:\n\n"
            "| Step | Required input | What you get |\n"
            "|---|---|---|\n"
            "| **MCR-ALS only** | Linescans + reference CSV *or* PCA init | Spatial component profiles, recovered pure spectra |\n"
            "| **PLS only** | Linescans + protein standard CSV | Quantitative protein (and optionally molecular crowder/salt) concentration profiles |\n"
            "| **MCR + PLS** | All of the above | Both, with spatial co-localisation scatter |\n"
            "| **Further analysis only** | Linescans (no standards needed) | PCA score plots, amide I decomposition, peak ratio profiles |\n\n"
            "**Nothing is mandatory except the linescan file(s).** "
            "**▶ Build PLS Model** is only needed if you want PLS concentration calibration. "
            "PLS predictions run automatically on all linescans immediately after building the model. "
            "MCR is fully independent — run it via **▶ Run MCR** in the MCR decomposition tab "
            "with or without a PLS model."
        )

    # ── Step-by-step guide ───────────────────────────────────────────────────
    with st.expander("Step-by-step guide", expanded=True):
        st.markdown(
            "#### 1 · Upload files  `📂 Files tab`\n"
            "- **Linescans** *(required)* — one or more `.txt` files exported from WITec software. "
            "The scan mode (lateral XY or depth Z) is **auto-detected** from the stage position "
            "data in each file — no manual selection needed.\n"
            "- **Protein standard CSV** *(optional)* — enables PLS protein quantification. "
            "Choose from the built-in training library or upload your own.\n"
            "- **Molecular crowder standard CSV** *(optional)* — when combined with the protein "
            "standard, a single dual-output PLS2 model predicts both concentrations simultaneously.\n"
            "- **Salt standard CSV** *(optional)* — trains a separate PLS model on the fingerprint "
            "region for ion concentration prediction.\n"
            "- **MCR reference CSV** *(optional)* — pure-component reference spectra used to "
            "initialise MCR-ALS. Alternatively, select **PCA from linescans** to estimate initial "
            "spectra automatically — no reference file needed.\n"
            "- **MCR components** — set the number of components (1–10) just below the MCR "
            "initialisation selector.\n\n"

            "#### 2 · Set spectral range  `📂 Files tab → Spectral range`\n"
            "- **Protein range** — wavenumber window used for PLS and MCR (default 700–3900 cm⁻¹). "
            "This setting does **not** affect Further analysis, which always uses the full raw range.\n"
            "- **Exclude gap region** *(toggle)* — remove a spectral gap from the model, e.g. the "
            "silent region 1850–2750 cm⁻¹.\n"
            "- **Salt range** — separate narrow window for the salt fingerprint band "
            "(default 900–1000 cm⁻¹).\n\n"

            "#### 3 · Set preprocessing  `📂 Files tab → Preprocessing`\n"
            "Each step can be switched or disabled independently:\n"
            "- **Baseline correction** — *Rubberband* (default), *Rolling ball*, *ALS*, or *None*.\n"
            "- **Normalisation** — *Min-max [0, 1]* (default), *SNV*, or *None*. "
            "Salt spectra have a separate normalisation setting (default: None).\n"
            "- **Spike removal** *(toggle)* — detects and replaces cosmic-ray spikes.\n\n"

            "#### 4 · Configure analysis settings  `📂 Files tab → Analysis settings`\n"
            "- **Concentration unit** and **protein MW** (for mM ↔ mg/mL conversion).\n"
            "- **MCR-ALS parameters** *(advanced)* — max iterations, convergence tolerance, "
            "regression method for C and Sᵀ (OLS or NNLS).\n"
            "- **PLS cross-validation** *(advanced)* — number of folds and max latent variables.\n\n"

            "#### 5 · Build PLS model  `📂 Files tab → ▶ Build PLS Model`  *(optional)*\n"
            "Trains PLS model(s) on the uploaded standard spectra. "
            "Once built, PLS predictions run automatically on all linescan files — results appear "
            "immediately in the **📊 PLS regression** tab. "
            "Skip this step if you only want MCR decomposition or further spectral analysis.\n\n"

            "#### 6 · Run MCR  `🔬 MCR decomposition tab → ▶ Run MCR`  *(optional)*\n"
            "Decomposes each linescan into spectral components using MCR-ALS. "
            "Requires either a reference CSV or PCA initialisation (configured in step 1). "
            "Results — concentration profiles **C** and recovered spectra **Sᵀ** — appear in the "
            "same tab. If PLS was also built, a co-localisation scatter plot is shown automatically.\n\n"

            "#### 7 · Further analysis  `🔍 Further analysis tab`  *(independent)*\n"
            "Available as soon as linescans have been processed (PLS/MCR not required). "
            "All sub-tools have their own spectral range and preprocessing controls, "
            "independent of the main pipeline. "
            "A dataset selector lets you include/exclude individual linescans. "
            "**Distance (µm)** and **Spectrum index** are always available as plot axis/colour "
            "options, alongside any PLS or MCR scores.\n\n"
            "- **PCA** — choose spectral region, normalisation (SNV / Min-max / None), and number of PCs. "
            "Score, loading, and explained-variance plots are shown.\n"
            "- **Amide I decomposition** — fits Gaussian components to the amide I band for each "
            "spectrum individually. Distributions of peak position, width, and area are reported.\n"
            "- **Peak ratio** — integrates two band windows and plots their ratio spatially and "
            "against any score label.\n\n"

            "#### 8 · Image overlay  `🗺️ Image overlay tab`\n"
            "Upload a microscopy image and map any score onto its spatial coordinates.\n"
            "- Select an objective magnification preset (100× / 20× / 5×) or enter custom dimensions.\n"
            "- **XY scan** — dots are placed at their true stage XY positions relative to the image "
            "centre you provide. Tick **Flip Y axis** if dots appear vertically mirrored.\n"
            "- **Z-scan** — image shows the acquisition position; a score-vs-depth plot appears alongside.\n"
            "- **Crop controls** — trim µm from any edge (left / right / top / bottom).\n\n"

            "#### 9 · Download  `⬇ Download tab`\n"
            "- **PLS / MCR results** — per-linescan Excel files or a bundled ZIP.\n"
            "- **Further analysis** — PCA scores/loadings/variance, amide fit parameters, "
            "and peak ratio profiles as separate downloads.\n"
            "- **Image overlay** — interactive HTML (zoomable in any browser) and static PDF "
            "(vector, publication-ready)."
        )

    # ── How the analysis works ───────────────────────────────────────────────
    with st.expander("How the analysis works — methods"):
        st.markdown(
            "##### Preprocessing pipeline\n"
            "Each spectrum passes through up to three steps (all individually selectable):\n\n"
            "1. **Baseline correction** — removes the broad fluorescence background. "
            "*Rubberband* fits a convex hull beneath the spectrum (via SpectroChemPy); "
            "*Rolling ball* uses morphological opening for curved baselines; "
            "*ALS* (Asymmetric Least Squares, via SpectroChemPy) fits a smooth baseline "
            "weighted towards the lower envelope. Choose *None* to skip.\n"
            "2. **Spike removal** — identifies cosmic-ray spikes using the SpectroChemPy "
            "despike algorithm and replaces them by interpolation from neighbouring points. "
            "Toggle off if your data contains genuine sharp features that should not be removed.\n"
            "3. **Normalisation** — *Min-max* scales each spectrum to [0, 1]; "
            "*SNV* (Standard Normal Variate) centres and scales to unit variance, "
            "correcting for multiplicative scatter. Choose *None* to preserve absolute "
            "intensities (recommended for salt PLS).\n\n"
            "A wavenumber gap cut can be toggled on to exclude the silent region "
            "(e.g. 1850–2750 cm⁻¹) before modelling. "
            "**Further analysis always uses the full raw spectral range** — the range "
            "set here does not affect it."
        )

        st.markdown(
            "##### Partial Least Squares (PLS) regression\n"
            "PLS builds latent variables (LVs) — linear combinations of wavenumbers — "
            "that simultaneously maximise spectral variance and covariance with the "
            "known concentration. A standard series is used for calibration; the model "
            "then predicts concentration at every position along the linescan.\n\n"
            "The number of LVs is chosen automatically by **k-fold cross-validation** "
            "(default k = 5, adjustable). To avoid selecting an overfitted model when "
            "the CV curve is flat, the **1-standard-error rule** is applied: the "
            "simplest model whose CV RMSE lies within one standard error of the "
            "minimum is preferred. The CV vs. training RMSE curve and the selected "
            "optimum are shown in the **📊 Calibration** tab.\n\n"
            "**Dual PLS (PLS2)** — when both a protein and a molecular crowder standard are uploaded, "
            "a single multi-output model is trained on the stacked dataset (protein "
            "spectra with crowder = 0, crowder spectra with protein = 0) and predicts both "
            "concentrations simultaneously.\n\n"
            "**Salt PLS** — a separate model trained on the fingerprint region "
            "(default 900–980 cm⁻¹) to quantify ion concentrations independently "
            "of the protein signal."
        )

        st.markdown(
            "##### Multivariate Curve Resolution – Alternating Least Squares (MCR-ALS)\n"
            "MCR-ALS decomposes the spectral matrix **X** into:\n\n"
            "&nbsp;&nbsp;&nbsp;&nbsp;**X ≈ C · Sᵀ**\n\n"
            "where **C** holds the concentration profiles and **Sᵀ** the pure-component "
            "spectra. The algorithm alternates between solving for **C** and **Sᵀ** with "
            "non-negativity constraints until convergence.\n\n"
            "**Initialisation** — two options, selectable in the Files tab:\n"
            "- *Reference CSV* — measured or literature reference spectra used as the "
            "initial **Sᵀ**. The reference spectra are preprocessed with the same pipeline "
            "as the linescans before being passed to MCR. Gives chemically interpretable "
            "components when good references are available.\n"
            "- *PCA from linescans* — PCA is computed on the linescan data and the "
            "non-negative-clipped loadings are used as the initial **Sᵀ**. No reference "
            "file needed; components are labelled comp\\_1, comp\\_2, … and represent "
            "dominant spectral patterns, not known pure species.\n\n"
            "MCR-ALS parameters (max iterations, convergence tolerance, regression method) "
            "are adjustable in the Files tab under Analysis settings."
        )

    # ── Further analysis ─────────────────────────────────────────────────────
    with st.expander("Further analysis — details"):
        st.markdown(
            "All three tools are in the **🔍 Further analysis** tab after running the "
            "main analysis. They re-read the raw linescan data and apply their own "
            "independent preprocessing — the spectral range and preprocessing set in "
            "the Files tab have no effect here. "
            "A **dataset selector** lets you include or exclude individual linescans. "
            "Concentration and MCR score range sliders (when available) filter spectra "
            "by phase. **Distance (µm)** and **Spectrum index** are always available "
            "as axis and colour options, regardless of whether PLS or MCR was run.\n\n"

            "**a. PCA of a selected spectral region**  \n"
            "A broad spectral cut and optional rubberband baseline are applied first; "
            "then a narrower analysis cut and a selectable normalisation "
            "(SNV / Min-max / None) are applied before PCA. "
            "The number of PCs is selectable (2–10). "
            "Score plots can be coloured by Distance, Spectrum index, PLS concentration, "
            "or MCR component score. Loadings and an explained-variance bar chart "
            "are shown alongside.\n\n"

            "**b. Amide I band decomposition**  \n"
            "Fits a user-defined number of Gaussians to the amide I region "
            "(default 1580–1720 cm⁻¹). Each spectrum is fitted individually, yielding "
            "distributions of peak centres, widths, and relative areas across the dataset. "
            "An individual spectrum viewer lets you inspect any single fit with an R² "
            "quality indicator. The mean spectrum is fitted separately and shown with a "
            "±1 SD band. Results are summarised as mean ± σ per component.\n\n"

            "**c. Peak ratio comparison**  \n"
            "Integrates two user-defined spectral windows and computes their ratio for "
            "every spectrum. The ratio is plotted as a spatial profile and as a scatter "
            "against any available score (Distance, Spectrum index, PLS, MCR, or PC)."
        )

    # ── File formats ─────────────────────────────────────────────────────────
    with st.expander("File formats"):
        st.markdown(
            "**Raw linescan files (`.txt`)**  \n"
            "Tab-delimited files exported from WITec instrument software. "
            "A header block contains the stage position coordinates (x, y, z in µm) "
            "for each spectrum; the data block has one spectrum per row, "
            "one wavenumber per column. "
            "Scan mode (XY or Z) is auto-detected from the position data — no manual "
            "selection needed.\n\n"
            "**PLS standard CSVs**  \n"
            "First column: concentration value (mg/mL, mM, or wt% — must match the unit "
            "selected in the Files tab). Remaining columns: raw spectral intensities at "
            "wavenumber points matching the linescan axis. "
            "The same preprocessing pipeline is applied to both standards and linescans "
            "before model fitting.\n\n"
            "**MCR reference CSV** *(optional)*  \n"
            "First column: a component label (e.g. Protein, Water, Glass). "
            "Remaining columns: reference spectrum intensities on the same wavenumber axis. "
            "Individual rows are selectable in the Files tab — include only the components "
            "relevant to your system. Not required when using PCA initialisation.\n\n"
            "**Microscopy image** *(Image overlay tab)*  \n"
            "Any standard format (.bmp, .png, .tif, .jpg). "
            "Provide the physical width and height (µm) of the field of view — "
            "magnification presets for 100×, 20×, and 5× are built in. "
            "An optional WITec position text file can be uploaded to set the image centre "
            "automatically from the stage coordinates recorded during acquisition."
        )


# ── Training data ──────────────────────────────────────────────────────────────
with tab_training:
    st.subheader("Training data")
    st.markdown(
        "**Spectral acquisition parameters (all standards)**\n\n"
        "| Parameter | Value |\n"
        "|---|---|\n"
        "| Objective | EC Epiplan-Neofluar 100×/0.9 DIC M27 (FWD = 1.0 mm) |\n"
        "| Laser wavelength | 532 nm |\n"
        "| Laser power | 10 mW |\n"
        "| Integration time | 60 s |\n"
        "| Accumulations | 1 |\n"
        "| Technical replicates | 5 |\n"
    )

    st.markdown("**Protein standards**")
    st.markdown(
        "| File | Protein | Supplier | CAS | Media | Source |\n"
        "|---|---|---|---|---|---|\n"
        "| `Insulin_10mW.csv` | Insulin, Human Recombinant | Sigma-Aldrich | 11061-68-0 | MQ pH 2 | Krog et al. *(in prep.)* |\n"
        "| `BSA_10mW.csv` | BSA | Sigma-Aldrich | 9048-46-8 | MQ | Krog et al. *(in prep.)* |\n"
        "| `LysC_10mW.csv` | Lysozyme (from hen egg whites) | Sigma-Aldrich | 12650-88-3 | MQ | Krog et al. *(in prep.)* |\n"
    )

    st.markdown("**Molecular crowder standards**")
    st.markdown(
        "| File | Grade | Supplier | CAS | Media | Source |\n"
        "|---|---|---|---|---|---|\n"
        "| `PEG4K_10mW.csv` | PEG 4000 (avg. MW 4000) | Sigma-Aldrich | 25322-68-3 | MQ | Krog et al. *(in prep.)* |\n"
        "| `PEG6K_10mW.csv` | PEG 6000 (avg. MW 6000) | Sigma-Aldrich | 25322-68-3 | MQ | Krog et al. *(in prep.)* |\n"
        "| `PEG8K_10mW.csv` | PEG 8000 (avg. MW 8000) | Sigma-Aldrich | 25322-68-3 | MQ | Krog et al. *(in prep.)* |\n"
    )

    st.markdown("**Salt standards**")
    st.markdown(
        "| File | Compound | Supplier | CAS | Media | Source |\n"
        "|---|---|---|---|---|---|\n"
        "| `SO4_10mW.csv` | Sodium sulfate (Na₂SO₄) | Sigma-Aldrich | 7757-82-6 | MQ pH 2 | Krog et al. *(in prep.)* |\n"
    )

    st.markdown("**MCR reference spectra**")
    st.markdown(
        "File: `MCR_reference_input_subtracted.csv`  \n"
        "Background-subtracted reference spectra used as initial guesses for MCR-ALS. "
        "Components are listed in document order (top to bottom) and can be selected "
        "individually via the ⚙️ Settings picker before building models.  \n"
        "Source: Krog et al. *(in prep.)*"
    )


# ── About ──────────────────────────────────────────────────────────────────────
with tab_about:
    st.subheader("About De-condensate")
    st.markdown("### Contact")
    st.markdown(
        "- Dr. Lasse Skjoldborg Krog† — [lasse.krog@sund.ku.dk](mailto:lasse.krog@sund.ku.dk)\n"
        "- Professor Vito Foderà† — [vito.fodera@sund.ku.dk](mailto:vito.fodera@sund.ku.dk)\n\n"
        "†Department of Pharmacy, University of Copenhagen, Universitetsparken 2, "
        "2100 Copenhagen, Denmark"
    )
    
    st.markdown("### Microscope")
    st.markdown(
        "Raman spectra are acquired on a "
        "[WITec alpha300 R confocal Raman microscope]"
        "(https://raman.oxinst.com/assets/uploads/raman/materials/WITec-alpha300-Brochure.pdf)."
    )

    st.markdown("### Python libraries & references")
    st.markdown(
        "| Library | Role | Reference |\n"
        "|---|---|---|\n"
        "| [scikit-learn](https://scikit-learn.org) · `PLSRegression` | PLS regression for concentration prediction | [Pedregosa et al., 2011](https://scikit-learn.org/stable/modules/generated/sklearn.cross_decomposition.PLSRegression.html) |\n"
        "| [pyMCR](https://pages.nist.gov/pyMCR/) | MCR-ALS spectral decomposition | [Camp, 2019](https://pages.nist.gov/pyMCR/) |\n"
        "| [SpectroChemPy](https://www.spectrochempy.fr) | Baseline correction (ALS, rubberband), Savitzky-Golay smoothing, spike removal | [Travert & Fernandez, J. Open Source Softw. 2023, 8(83), 5338](https://doi.org/10.21105/joss.05338) |\n"
        "| [NumPy](https://numpy.org) | Numerical array operations | Harris et al., 2020 |\n"
        "| [pandas](https://pandas.pydata.org) | Data handling and Excel export | McKinney, 2010 |\n"
        "| [Plotly](https://plotly.com/python/) | Interactive figures | Plotly Technologies Inc. |\n"
        "| [Streamlit](https://streamlit.io) | Web application framework | Streamlit Inc. |\n"
    )

